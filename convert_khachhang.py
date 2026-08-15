# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import test_cases from the original 83-case file
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from generate_testcase_excel_vn import test_cases

new_test_cases = []

# Process the original 83 test cases
for tc in test_cases:
    mod = tc["module"]
    title = tc["title"]
    steps = tc["steps"]
    
    # Determine platform
    platform = "App"
    if "CMS" in mod or "Admin" in title or "CMS" in steps or "CMS" in title:
        platform = "CMS"
        
    # Edge cases specifically checking App display after CMS action
    if "Hiển thị" in title and "App" in title:
        platform = "App"
        
    tc["platform"] = platform
    
    # Split some combined cases
    if "Duyệt TK Chờ duyệt" in title and "CMS" in platform:
        # Clone for App verification
        tc_app = dict(tc)
        tc_app["id"] = tc["id"] + "_App"
        tc_app["platform"] = "App"
        tc_app["title"] = "Kiểm tra App sau khi TK Chờ duyệt được Duyệt"
        tc_app["steps"] = "1. Mở App VINAGO\n2. Đăng nhập bằng TK vừa được duyệt"
        tc_app["expected"] = "- Đăng nhập thành công\n- Hiển thị đầy đủ tất cả Tab"
        
        tc["expected"] = "- Trạng thái TK -> Đang hoạt động\n- Hệ thống tự sinh Mã KH và Nickname\n- Hệ thống tự tạo mới thông tin Người mua"
        new_test_cases.append(tc)
        new_test_cases.append(tc_app)
    elif "TK phụ Đang hoạt động – áp dụng ngay trên App" in title:
        tc_app = dict(tc)
        tc_app["id"] = tc["id"] + "_App"
        tc_app["platform"] = "App"
        tc_app["title"] = "Kiểm tra quyền TK phụ Đang hoạt động trên App"
        tc_app["steps"] = "1. Mở App bằng TK phụ\n2. Kiểm tra giá bán"
        tc_app["expected"] = "- TK phụ ngay lập tức không thấy giá bán trên App"
        
        tc["title"] = "Điều chỉnh quyền TK phụ Đang hoạt động trên CMS"
        tc["expected"] = "- Thay đổi quyền lưu thành công trên CMS"
        new_test_cases.append(tc)
        new_test_cases.append(tc_app)
    else:
        new_test_cases.append(tc)

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    COLOR_HEADER      = "1F3864"
    COLOR_GROUP_ROW   = "2E75B6"
    COLOR_HAPPY       = "E2EFDA"
    COLOR_NEGATIVE    = "FCE4D6"
    COLOR_BOUNDARY    = "FFF2CC"
    COLOR_HIGH        = "C00000"
    COLOR_MEDIUM      = "ED7D31"
    COLOR_LOW         = "70AD47"
    COLOR_APP         = "D9EAD3" 
    COLOR_CMS         = "FFF2CC" 

    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_group   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_normal  = Font(name="Calibri", size=9)
    font_id      = Font(name="Calibri", bold=True, size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_left_c = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="595959")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "BỘ TEST CASE – APP VINAGO (MODULE KHÁCH HÀNG) (Cập nhật Tách App/CMS)"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 30

    headers = [
        "Mã TC", "Chức năng", "Test trên", "Tiêu đề", "Loại kiểm thử",
        "Điều kiện", "Các bước", "Kết quả mong đợi", "Ưu tiên"
    ]
    col_widths = [12, 16, 12, 35, 14, 30, 40, 40, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    groups = [
        ("NHÓM 1: ĐĂNG KÝ TÀI KHOẢN", "TC001", "TC019"),
        ("NHÓM 2: QUẢN LÝ TK", "TC020", "TC031"),
        ("NHÓM 3: TÌM KIẾM & BỘ LỌC", "TC032", "TC040"),
        ("NHÓM 4: NGƯỜI MUA", "TC041", "TC050"),
        ("NHÓM 5: TK CHÍNH PHỤ", "TC051", "TC060"),
        ("NHÓM 6: CẤP BẬC", "TC061", "TC068"),
        ("NHÓM 7: ĐỔI LOẠI KH", "TC069", "TC075"),
        ("NHÓM 8: HIỂN THỊ APP", "TC076", "TC083"),
    ]

    type_color = {"Luồng chuẩn": COLOR_HAPPY, "Luồng ngoại lệ": COLOR_NEGATIVE, "Giá trị biên": COLOR_BOUNDARY}
    priority_color = {"Cao": COLOR_HIGH, "Trung bình": COLOR_MEDIUM, "Thấp": COLOR_LOW}

    row_cursor = 3
    # Extract original base ids for group bounds
    tc_ids = [tc["id"].split("_")[0] for tc in new_test_cases]

    for group_name, start_id, end_id in groups:
        ws.merge_cells(f"A{row_cursor}:I{row_cursor}")
        g_cell = ws.cell(row=row_cursor, column=1, value=f"  {group_name}")
        g_cell.font, g_cell.fill, g_cell.alignment, g_cell.border = font_group, PatternFill("solid", fgColor=COLOR_GROUP_ROW), align_left_c, border_thick
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        start_idx = tc_ids.index(start_id)
        # find the last index that belongs to end_id
        end_idx = len(tc_ids) - 1 - tc_ids[::-1].index(end_id)
        
        for tc in new_test_cases[start_idx:end_idx+1]:
            bg = type_color.get(tc["type"], "FFFFFF")
            values = [tc["id"], tc["module"], tc["platform"], tc["title"], tc["type"], tc["preconditions"], tc["steps"], tc["expected"], tc["priority"]]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.fill, cell.border = PatternFill("solid", fgColor=bg), border_all
                if col_idx == 1: cell.font, cell.alignment = font_id, align_center
                elif col_idx == 3: 
                    plt = tc["platform"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("274E13" if plt == "App" else "B45F06"))
                    cell.fill = PatternFill("solid", fgColor=(COLOR_APP if plt == "App" else COLOR_CMS))
                    cell.alignment = align_center
                elif col_idx == 5:
                    c_txt = tc["type"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("276221" if c_txt=="Luồng chuẩn" else "7B2C00" if c_txt=="Luồng ngoại lệ" else "7B6300"))
                    cell.alignment = align_center
                elif col_idx == 9:
                    cell.font, cell.fill, cell.alignment = Font(name="Calibri", bold=True, color="FFFFFF", size=9), PatternFill("solid", fgColor=priority_color.get(tc.get("priority", "Trung bình"), "000000")), align_center
                elif col_idx in (2, 4): cell.font, cell.alignment = font_normal, align_left_c
                else: cell.font, cell.alignment = font_normal, align_left
            ws.row_dimensions[row_cursor].height = 60
            row_cursor += 1
        row_cursor += 1 

    ws2 = wb.create_sheet("Tổng kết")
    summary_headers = ["Nhóm", "Số TC", "Test trên CMS", "Test trên App"]
    summary_data = []
    
    for group_name, start_id, end_id in groups:
        s_idx = tc_ids.index(start_id)
        e_idx = len(tc_ids) - 1 - tc_ids[::-1].index(end_id)
        tcs = new_test_cases[s_idx:e_idx+1]
        summary_data.append((group_name, len(tcs), sum(1 for t in tcs if t["platform"] == "CMS"), sum(1 for t in tcs if t["platform"] == "App")))

    ws2.merge_cells("A1:D1")
    s_title = ws2["A1"]
    s_title.value, s_title.font, s_title.fill, s_title.alignment = "TỔNG KẾT", Font(name="Calibri", bold=True, color="FFFFFF", size=13), PatternFill("solid", fgColor="1F3864"), align_center
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = font_header, PatternFill("solid", fgColor=COLOR_HEADER), align_center, border_all
    
    for i, w in enumerate([30, 15, 15, 15], 1): ws2.column_dimensions[get_column_letter(i)].width = w

    totals = [0, 0, 0]
    for r_idx, row in enumerate(summary_data, 3):
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.font, cell.fill, cell.alignment, cell.border = Font(name="Calibri", size=10, bold=(col_idx==1)), PatternFill("solid", fgColor=("DEEAF1" if r_idx%2 else "FFFFFF")), align_center if col_idx > 1 else align_left_c, border_all
        totals[0] += row[1]; totals[1] += row[2]; totals[2] += row[3]

    output_path = r"d:\Java lean\TestCase\TestCase_VINAGO_KhachHang_TiengViet_Full.xlsx"
    wb.save(output_path)
    print("TC_KH_FULL_OK")

if __name__ == "__main__":
    create_excel()
