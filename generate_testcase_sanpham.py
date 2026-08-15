# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

test_cases = [
    # ── NHÓM 1: CẤU TRÚC DANH MỤC & SẢN PHẨM (APP & CMS) ─────────────────────────
    {
        "id": "TC_SP_001", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "Hiển thị đúng 4 cấp danh mục theo cấu trúc chuẩn",
        "type": "Luồng chuẩn",
        "preconditions": "Hệ thống đã đồng bộ cây thư mục từ Bravo. Có đủ 4 cấp.",
        "steps": "1. Mở App VINAGO -> Trang Sản phẩm\n2. Quan sát Left menu (Cấp 1)\n3. Chọn mở 1 Danh mục cấp 1\n4. Quan sát danh mục con bên phải (Cấp 2)\n5. Quan sát danh sách SKU (Cấp 4)",
        "expected": "- Cấp 1 hiển thị ở Left menu.\n- Cấp 2 hiển thị bên phải, dưới cấp 1 tương ứng.\n- Cấp 3 (Nhóm hàng) KHÔNG hiển thị trên App.\n- Cấp 4 (SKU) hiển thị danh sách sản phẩm.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_002", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "Thứ tự hiển thị mặc định của Danh mục Cấp 1 & Cấp 2",
        "type": "Luồng chuẩn",
        "preconditions": "CMS chưa có bất kỳ điều chỉnh thứ tự nào.",
        "steps": "1. Mở App VINAGO -> Trang Sản phẩm\n2. Quan sát thứ tự Danh mục Cấp 1 từ trên xuống\n3. Quan sát thứ tự Danh mục Cấp 2 trong 1 cấp 1",
        "expected": "- Thứ tự Cấp 1 và Cấp 2 hiển thị chính xác theo thứ tự mặc định của Bravo.",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_003", "module": "Cấu trúc Danh mục", "platform": "CMS",
        "title": "Thao tác sắp xếp lại vị trí Danh mục Cấp 2 trên CMS",
        "type": "Luồng chuẩn",
        "preconditions": "Đăng nhập CMS quyền Admin.",
        "steps": "1. CMS -> Danh mục hiển thị VINAGO\n2. Kéo thả/đổi vị trí Danh mục Cấp 2 (VD: đưa PLTTH 2 lên trước PLTTH 1)\n3. Lưu lại",
        "expected": "- CMS lưu thành công, hiển thị thông báo cập nhật vị trí.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_004", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "Kiểm tra vị trí Danh mục Cấp 2 trên App sau khi sắp xếp",
        "type": "Luồng chuẩn",
        "preconditions": "Admin đã sắp xếp vị trí PLTTH 2 lên trước PLTTH 1 trên CMS.",
        "steps": "1. Mở App VINAGO -> Trang Sản phẩm\n2. Chọn Danh mục Cấp 1 tương ứng",
        "expected": "- App hiển thị đúng vị trí Danh mục Cấp 2 mới cập nhật (PLTTH 2 nằm trên PLTTH 1).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_005", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "[Giá trị biên] Phát sinh Danh mục Cấp 2 mới khi đã điều chỉnh vị trí (App)",
        "type": "Giá trị biên",
        "preconditions": "Danh mục Cấp 1 đã được điều chỉnh vị trí các Cấp 2 con bên trong. Đồng bộ 1 Danh mục Cấp 2 mới từ Bravo.",
        "steps": "1. Mở App VINAGO -> Trang Sản phẩm\n2. Kiểm tra vị trí của Danh mục Cấp 2 mới",
        "expected": "- Danh mục Cấp 2 mới sẽ được đẩy lên hiển thị ở vị trí TRÊN CÙNG của Danh mục Cấp 1 đó.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_006", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "[Giá trị biên] Phát sinh Danh mục Cấp 2 mới khi CHƯA điều chỉnh vị trí (App)",
        "type": "Giá trị biên",
        "preconditions": "Danh mục Cấp 1 chưa từng điều chỉnh vị trí Cấp 2. Đồng bộ 1 Danh mục Cấp 2 mới từ Bravo.",
        "steps": "1. Mở App VINAGO -> Trang Sản phẩm\n2. Kiểm tra vị trí",
        "expected": "- Danh mục Cấp 2 mới hiển thị theo đúng thứ tự mặc định từ Bravo (thường là nối tiếp dưới cùng).",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_007", "module": "Cấu trúc Danh mục", "platform": "App",
        "title": "Đóng/Mở Danh mục Cấp 1 & Khảo sát",
        "type": "Luồng chuẩn",
        "preconditions": "User chưa thực hiện khảo sát cho Danh mục A. Danh mục A có gắn bộ câu hỏi.",
        "steps": "1. App -> Trang Sản phẩm\n2. Bấm chọn (Mở) Danh mục A trên Left menu",
        "expected": "- Hiển thị popup Bộ câu hỏi khảo sát của Danh mục A.\n- Bắt buộc trả lời khảo sát mới xem được sản phẩm.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_008", "module": "Chi tiết Sản phẩm", "platform": "CMS",
        "title": "Thiết lập loại giá cho sản phẩm trên CMS",
        "type": "Luồng chuẩn",
        "preconditions": "Đăng nhập CMS quyền Admin.",
        "steps": "1. Vào chi tiết 2 Sản phẩm (SP1, SP2)\n2. Thiết lập SP1: Loại giá = Giá liên hệ\n3. Thiết lập SP2: Loại giá = Giá định sẵn\n4. Lưu lại",
        "expected": "- Hệ thống lưu thành công thông tin loại giá cho từng sản phẩm.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_009", "module": "Chi tiết Sản phẩm", "platform": "App",
        "title": "Kiểm tra hiển thị loại giá (Liên hệ vs Định sẵn) trên App",
        "type": "Luồng chuẩn",
        "preconditions": "SP1 set Giá liên hệ, SP2 set Giá định sẵn.",
        "steps": "1. App -> Trang Sản phẩm\n2. Xem chi tiết SP1 và SP2",
        "expected": "- SP1 hiển thị chữ 'Liên hệ' thay vì mức giá cụ thể.\n- SP2 hiển thị đúng giá bán từ Bảng giá (áp dụng SL 1/SKU).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_010", "module": "Chi tiết Sản phẩm", "platform": "CMS",
        "title": "[Giá trị biên] Thiết lập tối đa 3 combo giá (CMS)",
        "type": "Giá trị biên",
        "preconditions": "Đăng nhập CMS.",
        "steps": "1. Vào chi tiết SP3\n2. Thêm 3 combo giá (Mua 2, Mua 5, Mua 10)\n3. Thử thêm combo giá thứ 4",
        "expected": "- Hệ thống không cho phép thêm combo thứ 4 (nút thêm bị mờ hoặc báo lỗi giới hạn Max 3).",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_011", "module": "Chi tiết Sản phẩm", "platform": "App",
        "title": "Hiển thị giá combo trên App",
        "type": "Luồng chuẩn",
        "preconditions": "CMS đã thiết lập SP3 có 3 combo giá.",
        "steps": "1. App -> Trang Sản phẩm -> Xem SP3",
        "expected": "- App hiển thị thông tin đúng 3 mức giá combo khi đạt số lượng yêu cầu.",
        "priority": "Trung bình"
    },

    # ── NHÓM 2: ẨN/HIỆN DANH MỤC SẢN PHẨM ─────────────────────────────────────
    {
        "id": "TC_SP_012", "module": "Ẩn/Hiện Danh mục", "platform": "App",
        "title": "Ẩn danh mục từ dữ liệu Gốc (TAFITCO) -> Ẩn luôn trên VINAGO",
        "type": "Luồng chuẩn",
        "preconditions": "Danh mục A thiết lập trạng thái Ẩn tại Danh mục Gốc (TAFITCO). CMS VINAGO Không cấu hình gì thêm.",
        "steps": "1. Mở App VINAGO kiểm tra Danh mục A.",
        "expected": "- Danh mục A không hiển thị trên App VINAGO (Ưu tiên Trạng thái gốc số 1).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_013", "module": "Ẩn/Hiện Danh mục", "platform": "CMS",
        "title": "Thiết lập ẩn Danh mục trên CMS VINAGO",
        "type": "Luồng chuẩn",
        "preconditions": "Danh mục B (Gốc = Hiện).",
        "steps": "1. CMS VINAGO -> Thiết lập Ẩn/Hiện danh mục\n2. Cài đặt Danh mục B = Ẩn\n3. Lưu",
        "expected": "- Cập nhật trạng thái thành công.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_014", "module": "Ẩn/Hiện Danh mục", "platform": "App",
        "title": "Kiểm tra ẩn danh mục trên App (khi thiết lập VINAGO = Ẩn)",
        "type": "Luồng chuẩn",
        "preconditions": "Danh mục B (Gốc = Hiện, Thiết lập VINAGO = Ẩn).",
        "steps": "1. Mở App VINAGO kiểm tra Danh mục B.",
        "expected": "- Danh mục B không hiển thị trên App VINAGO.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_015", "module": "Ẩn/Hiện Danh mục", "platform": "App",
        "title": "Thứ tự ưu tiên từ Cấp 1 -> Cấp 4 (Ẩn cấp cha -> Cấp con tự ẩn)",
        "type": "Luồng chuẩn",
        "preconditions": "Danh mục Cấp 1 bị Ẩn trên CMS VINAGO.",
        "steps": "1. App VINAGO -> Cố gắng tìm kiếm sản phẩm thuộc Cấp 1 đó.",
        "expected": "- Toàn bộ Cấp 2, Cấp 3, Cấp 4 (SKU) thuộc Cấp 1 đó đều bị ẩn và không xuất hiện trong kết quả tìm kiếm.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_016", "module": "Ẩn/Hiện Danh mục", "platform": "App",
        "title": "SKU (Cấp 4) ẩn theo Trạng thái bán hàng",
        "type": "Luồng chuẩn",
        "preconditions": "Danh mục Cấp 1, 2, 3 đều Hiện. SKU1 có Trạng thái bán hàng = Ẩn.",
        "steps": "1. App VINAGO -> Vào đúng danh mục chứa SKU1.",
        "expected": "- SKU1 không hiển thị trong danh sách sản phẩm.",
        "priority": "Cao"
    },

    # ── NHÓM 3: ẨN/HIỆN THEO ĐỐI TƯỢNG (PHIẾU THIẾT LẬP) ──────────────────────
    {
        "id": "TC_SP_017", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Tạo mới thiết lập (Trạng thái Nháp)",
        "type": "Luồng chuẩn",
        "preconditions": "CMS quyền Admin.",
        "steps": "1. CMS -> Ẩn hiện SP theo đối tượng -> Tạo mới\n2. Nhập Tên thiết lập, Chọn Đối tượng (VD: Tỉnh/Thành HCM)\n3. Xác nhận",
        "expected": "- Thiết lập tạo thành công với trạng thái 'Nháp'.\n- Có thể Kích hoạt, Xóa, hoặc Điều chỉnh.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_018", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Cảnh báo khi tạo trùng đối tượng",
        "type": "Luồng ngoại lệ",
        "preconditions": "Đã có thiết lập đang áp dụng cho Tỉnh HCM.",
        "steps": "1. CMS -> Tạo mới thiết lập -> Chọn Đối tượng = Tỉnh HCM.\n2. Bấm Xác nhận",
        "expected": "- Hiển thị cảnh báo: Đối tượng đã có thiết lập trước đó.",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_019", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Kích hoạt thiết lập & Tạo Phiếu áp dụng ngay",
        "type": "Luồng chuẩn",
        "preconditions": "Thiết lập đang Nháp.",
        "steps": "1. Kích hoạt thiết lập chính\n2. Tạo Phiếu cập nhật -> Thêm SP A (Trạng thái = Ẩn), Ngày áp dụng = Hiện tại\n3. Duyệt phiếu",
        "expected": "- Phiếu chuyển ngay sang trạng thái 'Đã áp dụng'.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_020", "module": "Ẩn/Hiện theo Đối tượng", "platform": "App",
        "title": "Kiểm tra SP bị Ẩn sau khi duyệt Phiếu áp dụng ngay",
        "type": "Luồng chuẩn",
        "preconditions": "Phiếu cập nhật SP A (Ẩn) cho Tỉnh HCM đã duyệt.",
        "steps": "1. App -> Đăng nhập tài khoản KH thuộc HCM\n2. Tìm kiếm SP A",
        "expected": "- SP A không hiển thị với khách hàng này.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_021", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Tạo phiếu cập nhật hẹn giờ áp dụng (Tương lai)",
        "type": "Luồng chuẩn",
        "preconditions": "Thiết lập đang hoạt động.",
        "steps": "1. Tạo Phiếu cập nhật -> Thêm SP B (Trạng thái = Hiện), Ngày áp dụng = Ngày mai\n2. Bấm Duyệt",
        "expected": "- Phiếu chuyển sang trạng thái 'Chờ áp dụng'.\n- (Tới đúng giờ ngày mai, phiếu tự động chuyển thành 'Đã áp dụng' mà không cần duyệt lại).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_022", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Hủy phiếu ở trạng thái Chờ áp dụng",
        "type": "Luồng chuẩn",
        "preconditions": "Phiếu cập nhật đang ở trạng thái 'Chờ áp dụng' (hẹn giờ tương lai).",
        "steps": "1. Chọn Phiếu 'Chờ áp dụng'\n2. Bấm Hủy",
        "expected": "- Phiếu chuyển sang trạng thái 'Hủy'. Hệ thống sẽ bỏ qua phiếu này khi đến giờ.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_023", "module": "Ẩn/Hiện theo Đối tượng", "platform": "App",
        "title": "Ưu tiên đối tượng: Chỉ định KH > Tỉnh/Thành",
        "type": "Logic ưu tiên",
        "preconditions": "Thiết lập 1: Tỉnh HCM -> Hiện SP A. Thiết lập 2: KH Nguyễn Văn B (thuộc HCM) -> Ẩn SP A.",
        "steps": "1. App -> Đăng nhập bằng tài khoản Nguyễn Văn B\n2. Tìm kiếm SP A",
        "expected": "- SP A BỊ ẨN đối với Nguyễn Văn B (vì thiết lập Chỉ định KH ưu tiên cao hơn Tỉnh/Thành).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_024", "module": "Ẩn/Hiện theo Đối tượng", "platform": "App",
        "title": "Ưu tiên Sản phẩm: Sản phẩm > Thương hiệu",
        "type": "Logic ưu tiên",
        "preconditions": "Thiết lập 1: Thương hiệu HYUNDAI -> Hiện. Thiết lập 2: SP Máy khoan X (thuộc HYUNDAI) -> Ẩn.",
        "steps": "1. App -> Xem danh sách SP của Thương hiệu HYUNDAI",
        "expected": "- Tất cả SP HYUNDAI đều hiển thị, NGOẠI TRỪ Máy khoan X bị ẩn (Ưu tiên Sản phẩm chi tiết cao hơn Thương hiệu).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_025", "module": "Ẩn/Hiện theo Đối tượng", "platform": "CMS",
        "title": "Xuất dữ liệu thiết lập Đang áp dụng",
        "type": "Luồng chuẩn",
        "preconditions": "CMS -> Thiết lập đang ở trạng thái Đang áp dụng.",
        "steps": "1. Bấm Xuất dữ liệu",
        "expected": "- File Excel được tải về thành công.\n- Đầy đủ cột: Mã SP, Tên SP, Nhóm hàng,... Trạng thái.",
        "priority": "Thấp"
    },

    # ── NHÓM 4: SẢN PHẨM BÁN CHẠY & SẢN PHẨM MỚI ───────────────────────────────
    {
        "id": "TC_SP_026", "module": "SP Bán Chạy", "platform": "App",
        "title": "Thứ tự ưu tiên hiển thị SP bán chạy trên App",
        "type": "Luồng chuẩn",
        "preconditions": "CMS có SP1 (Chỉ định thủ công), SP2 (Tần suất bán cao), SP3 (SL bán nhiều).",
        "steps": "1. App -> Trang Chủ (Khu vực SP bán chạy)",
        "expected": "- Thứ tự hiển thị: SP1 (Chỉ định) đứng đầu -> SP2 (Tần suất cao) -> SP3 (SL bán nhiều).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_027", "module": "SP Bán Chạy", "platform": "App",
        "title": "Giao diện lướt dọc & Thông tin SP Bán chạy",
        "type": "Luồng chuẩn",
        "preconditions": "Có 10 SP Bán chạy.",
        "steps": "1. App -> Khu vực SP bán chạy\n2. Vuốt xem danh sách\n3. Bấm Xem tất cả",
        "expected": "- Màn hình Home: Thông tin trái, giá phải, LƯỚT DỌC.\n- Bấm Xem tất cả: Chuyển sang hiển thị dạng lưới 2 cột bình thường.",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_028", "module": "SP Mới", "platform": "CMS",
        "title": "Import danh sách SP Mới bằng Excel (CMS)",
        "type": "Luồng chuẩn",
        "preconditions": "File excel chứa Mã SP hợp lệ, Ngày cập nhật = Ngày tương lai/Hiện tại.",
        "steps": "1. CMS -> Sản phẩm mới -> Import\n2. Upload file Excel\n3. Bấm Import",
        "expected": "- Import thành công.\n- SP mới hiển thị trong danh sách trên CMS.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_029", "module": "SP Mới", "platform": "CMS",
        "title": "Cảnh báo khi Import SP Mới với Ngày cập nhật là quá khứ",
        "type": "Luồng ngoại lệ",
        "preconditions": "File excel chứa Ngày cập nhật = Ngày hôm qua.",
        "steps": "1. CMS -> Sản phẩm mới -> Import file Excel",
        "expected": "- Báo lỗi: Ngày cập nhật không được là ngày trong quá khứ.\n- Dòng dữ liệu đó không được import.",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_030", "module": "SP Mới", "platform": "App",
        "title": "Giao diện hiển thị SP Mới trên App",
        "type": "Luồng chuẩn",
        "preconditions": "Có danh sách SP Mới đang kích hoạt.",
        "steps": "1. App -> Trang Chủ (Khu vực SP mới)",
        "expected": "- Ảnh/Thông tin ở trên, Giá ở dưới, LƯỚT NGANG.\n- Chỉ hiện Giá và Giảm giá (nếu có).\n- KHÔNG hiển thị Số lượng đã bán và Đánh giá sao.",
        "priority": "Cao"
    },

    # ── NHÓM 5: SẢN PHẨM TƯƠNG QUAN & KHÔNG TÍNH HOA HỒNG ──────────────────────
    {
        "id": "TC_SP_031", "module": "Sản Phẩm Tương Quan", "platform": "App",
        "title": "Tự động mở giá Danh mục Phụ tùng khi mua Máy",
        "type": "Luồng chuẩn",
        "preconditions": "Thiết lập: Mua 'Máy cắt cỏ 2 thì' (A) -> Tự mở giá 'Phụ tùng máy cắt cỏ' (B). KH chưa có giá B.",
        "steps": "1. KH trên App mua thành công 'Máy cắt cỏ 2 thì' (A).\n2. KH vào danh mục 'Phụ tùng máy cắt cỏ' (B).",
        "expected": "- Hệ thống tự động hiển thị giá (mở giá) cho Danh mục Phụ tùng (B) mà không cần Admin duyệt.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_032", "module": "Sản Phẩm Tương Quan", "platform": "CMS",
        "title": "[Giá trị biên] Đơn hàng từ CMS không trigger auto-mở giá",
        "type": "Giá trị biên",
        "preconditions": "Thiết lập A -> B. KH chưa có giá B.",
        "steps": "1. Admin tạo đơn hàng 'Máy cắt cỏ 2 thì' (A) cho KH trên CMS.\n2. Kiểm tra log/KH đăng nhập App xem danh mục B.",
        "expected": "- Danh mục B VẪN KHÔNG được mở giá (Rule: Không xét đơn tạo từ CMS).",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_033", "module": "Không tính Hoa hồng", "platform": "CMS",
        "title": "SP không tính HH vẫn tính vào tích lũy hạn mức, nhưng HH = 0",
        "type": "Logic hệ thống",
        "preconditions": "SP C nằm trong danh sách Không tính hoa hồng. KH đạt điều kiện thưởng HH 5%.",
        "steps": "1. KH mua đơn gồm SP D (10tr) và SP C (5tr)\n2. Chạy tính HH cuối kỳ trên CMS",
        "expected": "- Tổng tích lũy để xét hạn mức = 15tr.\n- Hoa hồng nhận được = 5% * 10tr (SP D). SP C không tính vào.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_034", "module": "Không tính Hoa hồng", "platform": "CMS",
        "title": "Hẹn giờ cập nhật danh sách Không tính hoa hồng",
        "type": "Luồng chuẩn",
        "preconditions": "CMS thiết lập Phiếu import danh sách SP không HH áp dụng vào ngày mai.",
        "steps": "1. Chờ tới ngày mai.\n2. Kiểm tra danh sách hiển thị trên CMS",
        "expected": "- Danh sách tự động cập nhật đúng ngày giờ hẹn.\n- Xóa danh sách cũ / thay bằng danh sách mới theo cấu hình.",
        "priority": "Trung bình"
    },

    # ── NHÓM 6: GIAO DIỆN APP (TAG, SỐ LƯỢNG, GIÁ) ─────────────────────────────
    {
        "id": "TC_SP_035", "module": "Giao diện App (Tag)", "platform": "App",
        "title": "Hiển thị Tag 'Cháy hàng' khi Tồn kho = 0",
        "type": "Luồng chuẩn",
        "preconditions": "SP X có tồn kho = 0 trong Bravo.",
        "steps": "1. App -> Trang Sản phẩm -> Tìm SP X",
        "expected": "- Hình ảnh/thông tin SP X hiển thị nhãn dán (Tag) 'Cháy hàng'.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_036", "module": "Giao diện App (Tag)", "platform": "App",
        "title": "Hiển thị Tag 'Sắp hết hàng' theo logic Max đơn",
        "type": "Luồng chuẩn",
        "preconditions": "Lịch sử xuất 30 ngày: Đơn 10, Đơn 50, Đơn 100. (Max = 100). Tồn kho khả dụng SP Y = 90.",
        "steps": "1. App -> Trang Sản phẩm -> Tìm SP Y",
        "expected": "- Hiển thị Tag 'Sắp hết hàng' vì Tồn (90) < Max xuất bán (100).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_037", "module": "Giao diện App (Tag)", "platform": "App",
        "title": "[Giá trị biên] Tag 'Sắp hết hàng' tự mất khi Tồn kho > Max đơn",
        "type": "Giá trị biên",
        "preconditions": "SP Y (Max = 100). Đã nhập thêm hàng, Tồn kho hiện tại = 150.",
        "steps": "1. App -> Tìm SP Y",
        "expected": "- Tag 'Sắp hết hàng' tự động biến mất.",
        "priority": "Trung bình"
    },
    {
        "id": "TC_SP_038", "module": "Giao diện App (Tag)", "platform": "App",
        "title": "SL bán hiển thị trên App không tính Hàng xuất tặng",
        "type": "Logic hệ thống",
        "preconditions": "Lịch sử Bravo: Xuất bán 50 cái, Xuất tặng 10 cái.",
        "steps": "1. App -> Kiểm tra SL đã bán của SP Z",
        "expected": "- Hiển thị Đã bán: 50 (Bỏ qua 10 cái xuất tặng).",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_039", "module": "Giao diện App (Giá)", "platform": "App",
        "title": "Hiển thị Giá Combo nếu thỏa số lượng n/SKU",
        "type": "Luồng chuẩn",
        "preconditions": "Combo: Mua 10 cái giảm còn 90k/cái. Giá gốc 100k.",
        "steps": "1. App -> Giỏ hàng -> Thêm 10 cái SP",
        "expected": "- Giỏ hàng tự động tính tổng tiền theo giá Combo 90k/cái.",
        "priority": "Cao"
    },
    {
        "id": "TC_SP_040", "module": "Giao diện App (Giá)", "platform": "App",
        "title": "Hiển thị Giá trước và sau giảm",
        "type": "Luồng chuẩn",
        "preconditions": "SP có tham gia CTKM giảm giá.",
        "steps": "1. App -> Trang Sản phẩm",
        "expected": "- Hiển thị Giá gốc bị gạch ngang (Giá trước).\n- Hiển thị Giá khuyến mãi nổi bật (Giá sau giảm).",
        "priority": "Cao"
    }
]


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    COLOR_HEADER      = "1F3864"
    COLOR_GROUP_ROW   = "2E75B6"
    COLOR_HAPPY       = "E2EFDA"
    COLOR_NEGATIVE    = "FCE4D6"
    COLOR_BOUNDARY    = "FFF2CC"
    COLOR_LOGIC       = "E6E6FA" 
    COLOR_HIGH        = "C00000"
    COLOR_MEDIUM      = "ED7D31"
    COLOR_LOW         = "70AD47"
    COLOR_APP         = "D9EAD3" # Xanh nhạt cho App
    COLOR_CMS         = "FFF2CC" # Vàng nhạt cho CMS

    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_group   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_normal  = Font(name="Calibri", size=9)
    font_id      = Font(name="Calibri", bold=True, size=9)
    font_platform = Font(name="Calibri", bold=True, size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_left_c = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="595959")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "BỘ TEST CASE – APP VINAGO (MODULE SẢN PHẨM)   |   Tài liệu: 260811_San pham - V1.pdf"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Mã Test Case", "Chức năng (Module)", "Test trên", "Tiêu đề (Mô tả)", "Loại kiểm thử",
        "Điều kiện tiên quyết", "Các bước thực hiện", "Kết quả mong đợi", "Mức độ ưu tiên"
    ]
    col_widths = [14, 25, 12, 42, 16, 38, 50, 45, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    groups = [
        ("NHÓM 1: CẤU TRÚC DANH MỤC & SẢN PHẨM (APP & CMS)", "TC_SP_001", "TC_SP_011"),
        ("NHÓM 2: ẨN/HIỆN DANH MỤC SẢN PHẨM",                "TC_SP_012", "TC_SP_016"),
        ("NHÓM 3: ẨN/HIỆN SẢN PHẨM THEO ĐỐI TƯỢNG",          "TC_SP_017", "TC_SP_025"),
        ("NHÓM 4: SẢN PHẨM BÁN CHẠY & SẢN PHẨM MỚI",          "TC_SP_026", "TC_SP_030"),
        ("NHÓM 5: SẢN PHẨM TƯƠNG QUAN & KHÔNG TÍNH HOA HỒNG", "TC_SP_031", "TC_SP_034"),
        ("NHÓM 6: GIAO DIỆN APP (TAG, SỐ LƯỢNG, GIÁ)",        "TC_SP_035", "TC_SP_040"),
    ]

    type_color = {
        "Luồng chuẩn":    COLOR_HAPPY,
        "Luồng ngoại lệ": COLOR_NEGATIVE,
        "Giá trị biên":   COLOR_BOUNDARY,
        "Logic hệ thống": COLOR_LOGIC,
    }
    priority_color = {
        "Cao":        COLOR_HIGH,
        "Trung bình": COLOR_MEDIUM,
        "Thấp":       COLOR_LOW,
    }

    row_cursor = 3
    tc_ids = [tc["id"] for tc in test_cases]

    for group_name, start_id, end_id in groups:
        ws.merge_cells(f"A{row_cursor}:I{row_cursor}")
        g_cell = ws.cell(row=row_cursor, column=1, value=f"  {group_name}")
        g_cell.font      = font_group
        g_cell.fill      = PatternFill("solid", fgColor=COLOR_GROUP_ROW)
        g_cell.alignment = Alignment(horizontal="left", vertical="center")
        g_cell.border    = border_thick
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        start_idx = tc_ids.index(start_id)
        end_idx   = tc_ids.index(end_id) + 1
        group_tcs = test_cases[start_idx:end_idx]

        for tc in group_tcs:
            bg = type_color.get(tc["type"], "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)

            values = [
                tc["id"], tc["module"], tc["platform"], tc["title"], tc["type"],
                tc["preconditions"], tc["steps"],
                tc["expected"], tc["priority"]
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.fill   = fill
                cell.border = border_all

                if col_idx == 1:  
                    cell.font      = font_id
                    cell.alignment = align_center
                elif col_idx == 3: 
                    platform = tc["platform"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("274E13" if platform == "App" else "B45F06"))
                    cell.fill = PatternFill("solid", fgColor=(COLOR_APP if platform == "App" else COLOR_CMS))
                    cell.alignment = align_center
                elif col_idx == 5:  
                    c_txt = tc["type"]
                    cell.font      = Font(name="Calibri", bold=True, size=9,
                                          color=("276221" if c_txt=="Luồng chuẩn"
                                                 else "7B2C00" if c_txt=="Luồng ngoại lệ"
                                                 else "4B0082" if c_txt=="Logic hệ thống"
                                                 else "7B6300"))
                    cell.alignment = align_center
                elif col_idx == 9:  
                    p_color = priority_color.get(tc["priority"], "000000")
                    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
                    cell.fill      = PatternFill("solid", fgColor=p_color)
                    cell.alignment = align_center
                elif col_idx in (2, 4):
                    cell.font      = font_normal
                    cell.alignment = align_left_c
                else:
                    cell.font      = font_normal
                    cell.alignment = align_left

            ws.row_dimensions[row_cursor].height = 80
            row_cursor += 1

        row_cursor += 1 

    ws2 = wb.create_sheet("Tổng kết")
    summary_headers = ["Nhóm chức năng", "Số Test Case", "Test trên CMS", "Test trên App", "Luồng chuẩn", "Luồng ngoại lệ", "Biên/Logic"]
    
    summary_data = []
    for group_name, start_id, end_id in groups:
        s_idx = tc_ids.index(start_id)
        e_idx = tc_ids.index(end_id) + 1
        group_tcs = test_cases[s_idx:e_idx]
        
        total = len(group_tcs)
        cms_cnt = sum(1 for tc in group_tcs if tc["platform"] == "CMS")
        app_cnt = sum(1 for tc in group_tcs if tc["platform"] == "App")
        hc_cnt = sum(1 for tc in group_tcs if tc["type"] == "Luồng chuẩn")
        nc_cnt = sum(1 for tc in group_tcs if tc["type"] == "Luồng ngoại lệ")
        bl_cnt = sum(1 for tc in group_tcs if tc["type"] in ["Giá trị biên", "Logic hệ thống"])
        
        short_name = group_name.split(":")[0] + ": " + group_name.split(":")[1].split("(")[0].strip()
        summary_data.append((short_name, total, cms_cnt, app_cnt, hc_cnt, nc_cnt, bl_cnt))

    ws2.merge_cells("A1:G1")
    s_title = ws2["A1"]
    s_title.value     = "TỔNG KẾT BỘ TEST CASE – MODULE SẢN PHẨM"
    s_title.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    s_title.fill      = PatternFill("solid", fgColor="1F3864")
    s_title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all

    s_widths = [45, 15, 15, 15, 15, 18, 15]
    for i, w in enumerate(s_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    totals = [0, 0, 0, 0, 0, 0]
    for r_idx, row in enumerate(summary_data, start=3):
        fills = [PatternFill("solid", fgColor="DEEAF1"), PatternFill("solid", fgColor="FFFFFF")]
        row_fill = fills[r_idx % 2]
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10, bold=(col_idx==1))
            cell.fill      = row_fill
            cell.alignment = align_center if col_idx > 1 else Alignment(horizontal="left", vertical="center")
            cell.border    = border_all
        
        for i in range(6):
            totals[i] += row[i+1]

    total_row = r_idx + 1
    cell_tong = ws2.cell(row=total_row, column=1, value="TỔNG CỘNG")
    cell_tong.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell_tong.fill      = PatternFill("solid", fgColor="1F3864")
    cell_tong.alignment = Alignment(horizontal="center", vertical="center")
    cell_tong.border    = border_thick

    for col_idx, val in enumerate(totals, start=2):
        cell = ws2.cell(row=total_row, column=col_idx, value=val)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = align_center
        cell.border    = border_thick

    ws2.row_dimensions[total_row].height = 22

    output_path = r"d:\Java lean\TestCase\TestCase_VINAGO_SanPham_TiengViet.xlsx"
    wb.save(output_path)
    print("TC_OK_NEW")

if __name__ == "__main__":
    create_excel()
