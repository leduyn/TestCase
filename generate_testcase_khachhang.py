# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

test_cases = [
    # NHÓM 1: ĐĂNG KÝ TÀI KHOẢN (APP)
    {"id": "TC_KH_001", "module": "Đăng ký TK", "platform": "App", "title": "Đăng ký thành công với đầy đủ thông tin hợp lệ", "type": "Luồng chuẩn", "preconditions": "App đang ở màn hình Đăng ký, SĐT và MST chưa tồn tại", "steps": "1. Nhập Họ tên, SĐT, Tên CH\n2. Chọn Tỉnh/Phường\n3. Nhập Địa chỉ, MST\n4. Chờ auto-fill từ Cổng Thuế\n5. Nhập MK, upload ảnh, chọn danh mục\n6. Trả lời khảo sát\n7. Nhấn Đăng ký", "expected": "- Đăng ký thành công\n- Trạng thái TK là Chờ duyệt trên CMS", "priority": "Cao"},
    {"id": "TC_KH_002", "module": "Đăng ký TK", "platform": "App", "title": "Thông tin XHĐ tự động lấy từ Cổng Thuế", "type": "Luồng chuẩn", "preconditions": "Màn hình Đăng ký", "steps": "1. Điền thông tin\n2. Nhập MST có thật\n3. Bấm ra ngoài", "expected": "- Tên đơn vị và Địa chỉ XHĐ tự động điền từ Cổng Thuế", "priority": "Cao"},
    {"id": "TC_KH_003", "module": "Đăng ký TK", "platform": "App", "title": "Cổng Thuế không trả về dữ liệu", "type": "Luồng chuẩn", "preconditions": "Màn hình Đăng ký", "steps": "1. Nhập MST không tồn tại\n2. Nhập thủ công tên đơn vị", "expected": "- Cho phép nhập thủ công\n- Tên tự động IN HOA", "priority": "Trung bình"},
    {"id": "TC_KH_004", "module": "Đăng ký TK", "platform": "App", "title": "Bỏ trống trường Họ và tên", "type": "Luồng ngoại lệ", "preconditions": "Màn hình Đăng ký", "steps": "1. Để trống Họ và tên\n2. Submit", "expected": "- Báo lỗi Vui lòng nhập Họ và tên", "priority": "Cao"},
    {"id": "TC_KH_005", "module": "Đăng ký TK", "platform": "App", "title": "Bỏ trống Số điện thoại", "type": "Luồng ngoại lệ", "preconditions": "Màn hình Đăng ký", "steps": "1. Để trống SĐT\n2. Submit", "expected": "- Báo lỗi Vui lòng nhập Số điện thoại", "priority": "Cao"},
    {"id": "TC_KH_006", "module": "Đăng ký TK", "platform": "App", "title": "SĐT đã tồn tại", "type": "Luồng ngoại lệ", "preconditions": "SĐT đã được đăng ký", "steps": "1. Nhập SĐT trùng\n2. Submit", "expected": "- Báo lỗi Số điện thoại đã được sử dụng", "priority": "Cao"},
    {"id": "TC_KH_007", "module": "Đăng ký TK", "platform": "App", "title": "MST đã tồn tại", "type": "Luồng ngoại lệ", "preconditions": "MST đã được đăng ký", "steps": "1. Nhập MST trùng\n2. Submit", "expected": "- Báo lỗi Mã số thuế đã được đăng ký", "priority": "Cao"},
    {"id": "TC_KH_008", "module": "Đăng ký TK", "platform": "App", "title": "MST trùng với TK đã bị Từ chối", "type": "Luồng chuẩn", "preconditions": "MST tồn tại nhưng trạng thái Từ chối", "steps": "1. Nhập MST\n2. Submit", "expected": "- Cho phép đăng ký thành công", "priority": "Trung bình"},
    {"id": "TC_KH_009", "module": "Đăng ký TK", "platform": "App", "title": "[Biên] Mật khẩu 5 ký tự", "type": "Giá trị biên", "preconditions": "Màn hình Đăng ký", "steps": "1. Nhập MK 5 ký tự\n2. Submit", "expected": "- Báo lỗi tối thiểu 6 ký tự", "priority": "Cao"},
    {"id": "TC_KH_010", "module": "Đăng ký TK", "platform": "App", "title": "Mật khẩu không khớp", "type": "Luồng ngoại lệ", "preconditions": "Màn hình Đăng ký", "steps": "1. Nhập MK và Nhập lại MK khác nhau", "expected": "- Báo lỗi Mật khẩu không khớp", "priority": "Cao"},
    {"id": "TC_KH_011", "module": "Đăng ký TK", "platform": "App", "title": "Không chọn danh mục sản phẩm", "type": "Luồng ngoại lệ", "preconditions": "Màn hình Đăng ký", "steps": "1. Không chọn danh mục\n2. Submit", "expected": "- Báo lỗi Vui lòng chọn danh mục", "priority": "Cao"},

    # NHÓM 2: QUẢN LÝ TRẠNG THÁI TÀI KHOẢN (CMS & APP)
    {"id": "TC_KH_012", "module": "Quản lý TK", "platform": "CMS", "title": "Duyệt TK Chờ duyệt", "type": "Luồng chuẩn", "preconditions": "TK trạng thái Chờ duyệt", "steps": "1. CMS -> DS TK\n2. Bấm Duyệt", "expected": "- Trạng thái -> Đang hoạt động\n- Sinh Mã KH và Nickname", "priority": "Cao"},
    {"id": "TC_KH_013", "module": "Quản lý TK", "platform": "App", "title": "Kiểm tra TK sau khi được Duyệt trên App", "type": "Luồng chuẩn", "preconditions": "TK đã được Admin duyệt", "steps": "1. Đăng nhập App", "expected": "- Đăng nhập thành công\n- Hiển thị đầy đủ Tab mua hàng", "priority": "Cao"},
    {"id": "TC_KH_014", "module": "Quản lý TK", "platform": "CMS", "title": "Từ chối TK Chờ duyệt", "type": "Luồng chuẩn", "preconditions": "TK Chờ duyệt", "steps": "1. CMS -> Bấm Từ chối", "expected": "- Trạng thái -> Từ chối", "priority": "Trung bình"},
    {"id": "TC_KH_015", "module": "Quản lý TK", "platform": "CMS", "title": "Tạm dừng TK Đang hoạt động", "type": "Luồng chuẩn", "preconditions": "TK Đang hoạt động", "steps": "1. CMS -> Bấm Tạm dừng", "expected": "- Trạng thái -> Ngưng hoạt động", "priority": "Cao"},
    {"id": "TC_KH_016", "module": "Quản lý TK", "platform": "App", "title": "Đăng nhập với TK Ngưng hoạt động", "type": "Luồng ngoại lệ", "preconditions": "TK Ngưng hoạt động", "steps": "1. Đăng nhập App", "expected": "- Báo lỗi tài khoản bị tạm dừng, không vào được App", "priority": "Cao"},
    {"id": "TC_KH_017", "module": "Quản lý TK", "platform": "CMS", "title": "Tạo mới TK từ CMS", "type": "Luồng chuẩn", "preconditions": "Admin CMS", "steps": "1. CMS -> Tạo mới\n2. Nhập MST và lưu", "expected": "- Tạo TK thành công, tự fill thuế", "priority": "Trung bình"},
    {"id": "TC_KH_018", "module": "Quản lý TK", "platform": "CMS", "title": "[Biên] Nickname 16 ký tự", "type": "Giá trị biên", "preconditions": "CMS", "steps": "1. Sửa Nickname thành 16 ký tự", "expected": "- Không cho phép lưu hoặc tự cắt", "priority": "Trung bình"},
    {"id": "TC_KH_019", "module": "Quản lý TK", "platform": "App", "title": "Sửa Nickname chứa ký tự đặc biệt", "type": "Luồng ngoại lệ", "preconditions": "App", "steps": "1. Sửa Nickname nhập @#$", "expected": "- Báo lỗi ký tự không hợp lệ", "priority": "Trung bình"},

    # NHÓM 3: DANH SÁCH NGƯỜI MUA (CMS)
    {"id": "TC_KH_020", "module": "DS Người mua", "platform": "CMS", "title": "Tạo mới Người mua", "type": "Luồng chuẩn", "preconditions": "Admin CMS", "steps": "1. CMS -> Người mua -> Tạo mới\n2. Nhập MST hợp lệ", "expected": "- Lưu thành công ở trạng thái Nháp", "priority": "Cao"},
    {"id": "TC_KH_021", "module": "DS Người mua", "platform": "CMS", "title": "Tạo mới Người mua với MST trùng", "type": "Luồng ngoại lệ", "preconditions": "MST đã tồn tại", "steps": "1. Tạo mới với MST cũ", "expected": "- Báo lỗi MST đã tồn tại", "priority": "Cao"},
    {"id": "TC_KH_022", "module": "DS Người mua", "platform": "CMS", "title": "Luồng duyệt Người mua: Nháp -> Đang giao dịch", "type": "Luồng chuẩn", "preconditions": "Người mua Nháp", "steps": "1. Xác nhận\n2. Duyệt", "expected": "- Trạng thái chuyển thành Đang giao dịch", "priority": "Cao"},
    {"id": "TC_KH_023", "module": "DS Người mua", "platform": "CMS", "title": "Tạm dừng Người mua", "type": "Luồng chuẩn", "preconditions": "Người mua Đang giao dịch", "steps": "1. Nhấn Tạm dừng", "expected": "- Trạng thái -> Ngưng giao dịch", "priority": "Trung bình"},

    # NHÓM 4: TÀI KHOẢN CHÍNH / PHỤ
    {"id": "TC_KH_024", "module": "TK Phụ", "platform": "App", "title": "Tạo TK phụ từ TK chính", "type": "Luồng chuẩn", "preconditions": "TK chính", "steps": "1. App -> Quản lý TK -> Tạo TK phụ\n2. Nhập thông tin, SĐT", "expected": "- Trạng thái Chờ kích hoạt\n- OTP gửi về TK chính", "priority": "Cao"},
    {"id": "TC_KH_025", "module": "TK Phụ", "platform": "CMS", "title": "Admin kích hoạt TK phụ", "type": "Luồng chuẩn", "preconditions": "TK phụ chờ kích hoạt", "steps": "1. CMS -> Kích hoạt", "expected": "- TK phụ chuyển thành Đang hoạt động", "priority": "Cao"},
    {"id": "TC_KH_026", "module": "TK Phụ", "platform": "App", "title": "Đăng nhập TK phụ sau khi kích hoạt", "type": "Luồng chuẩn", "preconditions": "TK phụ đang hoạt động", "steps": "1. Đăng nhập App", "expected": "- Đăng nhập thành công với quyền mặc định", "priority": "Cao"},
    {"id": "TC_KH_027", "module": "TK Phụ", "platform": "CMS", "title": "Điều chỉnh quyền Không thấy công nợ", "type": "Luồng chuẩn", "preconditions": "Admin CMS", "steps": "1. Tắt quyền xem công nợ của TK phụ", "expected": "- Lưu thành công", "priority": "Trung bình"},
    {"id": "TC_KH_028", "module": "TK Phụ", "platform": "App", "title": "Kiểm tra quyền Không thấy công nợ trên App", "type": "Luồng ngoại lệ", "preconditions": "TK phụ bị tắt quyền", "steps": "1. Mở App bằng TK phụ\n2. Tìm mục Công nợ", "expected": "- Mục công nợ bị ẩn/mờ", "priority": "Cao"},
    {"id": "TC_KH_029", "module": "TK Phụ", "platform": "App", "title": "Quên MK TK phụ gửi OTP về TK chính", "type": "Luồng chuẩn", "preconditions": "Màn hình quên MK", "steps": "1. Nhập SĐT TK phụ", "expected": "- OTP gửi về số của TK chính", "priority": "Cao"},

    # NHÓM 5: HỆ THỐNG CẤP BẬC
    {"id": "TC_KH_030", "module": "Cấp bậc", "platform": "CMS", "title": "Cấp bậc mặc định cho Sỉ", "type": "Luồng chuẩn", "preconditions": "Duyệt KH Sỉ mới", "steps": "1. Admin duyệt", "expected": "- Cấp Thạch Anh", "priority": "Cao"},
    {"id": "TC_KH_031", "module": "Cấp bậc", "platform": "App", "title": "Hiển thị cấp bậc cho KH Sỉ", "type": "Luồng chuẩn", "preconditions": "KH Sỉ Hổ Phách", "steps": "1. App -> Cấp bậc", "expected": "- Hiển thị cấp hiện tại (Hổ Phách) và kế tiếp (Ngọc Bích)", "priority": "Trung bình"},
    {"id": "TC_KH_032", "module": "Cấp bậc", "platform": "App", "title": "KH Lẻ không có cấp bậc", "type": "Luồng ngoại lệ", "preconditions": "KH Lẻ", "steps": "1. App -> Cấp bậc", "expected": "- Hiển thị Thành viên, không có rank", "priority": "Trung bình"},
    {"id": "TC_KH_033", "module": "Cấp bậc", "platform": "CMS", "title": "[Biên] Thăng cấp khi đủ DTT và Tiền thu", "type": "Giá trị biên", "preconditions": "Cuối tháng, KH đủ điều kiện", "steps": "1. Chạy job cập nhật", "expected": "- KH lên cấp mới", "priority": "Cao"},

    # NHÓM 6: CHUYỂN ĐỔI LOẠI KH
    {"id": "TC_KH_034", "module": "Đổi Loại KH", "platform": "CMS", "title": "Lẻ -> Sỉ: Cấp bậc về Thạch Anh", "type": "Luồng chuẩn", "preconditions": "KH Lẻ", "steps": "1. CMS -> Đổi sang Sỉ", "expected": "- Lưu loại Sỉ, cấp Thạch Anh", "priority": "Cao"},
    {"id": "TC_KH_035", "module": "Đổi Loại KH", "platform": "CMS", "title": "Sỉ -> Lẻ: Cấn trừ phí bảo lãnh", "type": "Luồng chuẩn", "preconditions": "KH Sỉ có phí bảo lãnh và công nợ", "steps": "1. CMS -> Đổi sang Lẻ", "expected": "- Cấp Thành viên\n- Phí bảo lãnh tự động cấn trừ công nợ", "priority": "Cao"},
    {"id": "TC_KH_036", "module": "Đổi Loại KH", "platform": "App", "title": "Kiểm tra hiển thị sau khi đổi Lẻ -> Sỉ", "type": "Luồng chuẩn", "preconditions": "KH đã đổi thành Sỉ", "steps": "1. Đăng nhập App", "expected": "- Hiển thị tab công nợ, cấp bậc Sỉ", "priority": "Cao"},
    {"id": "TC_KH_037", "module": "Đổi Loại KH", "platform": "App", "title": "Sỉ -> Lẻ: Không chọn được XHĐ người khác", "type": "Giá trị biên", "preconditions": "KH đã thành Lẻ", "steps": "1. App -> Đặt hàng -> Chọn XHĐ", "expected": "- Chỉ thấy XHĐ của chính mình", "priority": "Trung bình"},
]

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    COLOR_HEADER      = "1F3864"
    COLOR_GROUP_ROW   = "2E75B6"
    COLOR_HAPPY       = "E2EFDA"
    COLOR_NEGATIVE    = "FCE4D6"
    COLOR_BOUNDARY    = "FFF2CC"
    COLOR_HIGH        = "C00000"
    COLOR_MEDIUM      = "ED7D31"
    COLOR_LOW         = "70AD47"
    COLOR_APP         = "D9EAD3" 
    COLOR_CMS         = "FFF2CC" 

    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_group   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_normal  = Font(name="Calibri", size=9)
    font_id      = Font(name="Calibri", bold=True, size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_left_c = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="595959")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "BỘ TEST CASE – APP VINAGO (MODULE KHÁCH HÀNG) (Cập nhật Tách App/CMS)"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 30

    headers = [
        "Mã TC", "Chức năng", "Test trên", "Tiêu đề", "Loại kiểm thử",
        "Điều kiện", "Các bước", "Kết quả mong đợi", "Ưu tiên"
    ]
    col_widths = [12, 16, 12, 35, 14, 30, 40, 40, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    groups = [
        ("NHÓM 1: ĐĂNG KÝ TÀI KHOẢN", "TC_KH_001", "TC_KH_011"),
        ("NHÓM 2: QUẢN LÝ TK", "TC_KH_012", "TC_KH_019"),
        ("NHÓM 3: NGƯỜI MUA", "TC_KH_020", "TC_KH_023"),
        ("NHÓM 4: TK CHÍNH PHỤ", "TC_KH_024", "TC_KH_029"),
        ("NHÓM 5: CẤP BẬC", "TC_KH_030", "TC_KH_033"),
        ("NHÓM 6: ĐỔI LOẠI KH", "TC_KH_034", "TC_KH_037"),
    ]

    type_color = {"Luồng chuẩn": COLOR_HAPPY, "Luồng ngoại lệ": COLOR_NEGATIVE, "Giá trị biên": COLOR_BOUNDARY}
    priority_color = {"Cao": COLOR_HIGH, "Trung bình": COLOR_MEDIUM, "Thấp": COLOR_LOW}

    row_cursor = 3
    tc_ids = [tc["id"] for tc in test_cases]

    for group_name, start_id, end_id in groups:
        ws.merge_cells(f"A{row_cursor}:I{row_cursor}")
        g_cell = ws.cell(row=row_cursor, column=1, value=f"  {group_name}")
        g_cell.font, g_cell.fill, g_cell.alignment, g_cell.border = font_group, PatternFill("solid", fgColor=COLOR_GROUP_ROW), align_left_c, border_thick
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        start_idx = tc_ids.index(start_id)
        end_idx   = tc_ids.index(end_id) + 1
        for tc in test_cases[start_idx:end_idx]:
            bg = type_color.get(tc["type"], "FFFFFF")
            values = [tc["id"], tc["module"], tc["platform"], tc["title"], tc["type"], tc["preconditions"], tc["steps"], tc["expected"], tc["priority"]]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.fill, cell.border = PatternFill("solid", fgColor=bg), border_all
                if col_idx == 1: cell.font, cell.alignment = font_id, align_center
                elif col_idx == 3: 
                    plt = tc["platform"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("274E13" if plt == "App" else "B45F06"))
                    cell.fill = PatternFill("solid", fgColor=(COLOR_APP if plt == "App" else COLOR_CMS))
                    cell.alignment = align_center
                elif col_idx == 5:
                    c_txt = tc["type"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("276221" if c_txt=="Luồng chuẩn" else "7B2C00" if c_txt=="Luồng ngoại lệ" else "7B6300"))
                    cell.alignment = align_center
                elif col_idx == 9:
                    cell.font, cell.fill, cell.alignment = Font(name="Calibri", bold=True, color="FFFFFF", size=9), PatternFill("solid", fgColor=priority_color.get(tc["priority"], "000000")), align_center
                elif col_idx in (2, 4): cell.font, cell.alignment = font_normal, align_left_c
                else: cell.font, cell.alignment = font_normal, align_left
            ws.row_dimensions[row_cursor].height = 60
            row_cursor += 1
        row_cursor += 1 

    ws2 = wb.create_sheet("Tổng kết")
    summary_headers = ["Nhóm", "Số TC", "Test trên CMS", "Test trên App"]
    summary_data = []
    for group_name, start_id, end_id in groups:
        s_idx, e_idx = tc_ids.index(start_id), tc_ids.index(end_id) + 1
        tcs = test_cases[s_idx:e_idx]
        summary_data.append((group_name, len(tcs), sum(1 for t in tcs if t["platform"] == "CMS"), sum(1 for t in tcs if t["platform"] == "App")))

    ws2.merge_cells("A1:D1")
    s_title = ws2["A1"]
    s_title.value, s_title.font, s_title.fill, s_title.alignment = "TỔNG KẾT", Font(name="Calibri", bold=True, color="FFFFFF", size=13), PatternFill("solid", fgColor="1F3864"), align_center
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = font_header, PatternFill("solid", fgColor=COLOR_HEADER), align_center, border_all
    
    for i, w in enumerate([30, 15, 15, 15], 1): ws2.column_dimensions[get_column_letter(i)].width = w

    totals = [0, 0, 0]
    for r_idx, row in enumerate(summary_data, 3):
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.font, cell.fill, cell.alignment, cell.border = Font(name="Calibri", size=10, bold=(col_idx==1)), PatternFill("solid", fgColor=("DEEAF1" if r_idx%2 else "FFFFFF")), align_center if col_idx > 1 else align_left_c, border_all
        totals[0] += row[1]; totals[1] += row[2]; totals[2] += row[3]

    output_path = r"d:\Java lean\TestCase\TestCase_VINAGO_KhachHang_TiengViet_Updated.xlsx"
    wb.save(output_path)
    print("TC_KH_OK")

if __name__ == "__main__":
    create_excel()
