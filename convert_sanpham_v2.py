# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from generate_testcase_sanpham import test_cases

for tc in test_cases:
    platform = tc.get("platform", "App")
    tc["server"] = "Staging"
    tc["device"] = "Mobile" if platform == "App" else "PC/Web"

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    COLOR_HEADER      = "1F3864"
    COLOR_GROUP_ROW   = "2E75B6"
    COLOR_HAPPY       = "E2EFDA"
    COLOR_NEGATIVE    = "FCE4D6"
    COLOR_BOUNDARY    = "FFF2CC"
    COLOR_LOGIC       = "E6E6FA" 
    COLOR_HIGH        = "C00000"
    COLOR_MEDIUM      = "ED7D31"
    COLOR_LOW         = "70AD47"
    COLOR_APP         = "D9EAD3" 
    COLOR_CMS         = "FFF2CC" 

    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_group   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_normal  = Font(name="Calibri", size=9)
    font_id      = Font(name="Calibri", bold=True, size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_left_c = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="595959")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "BỘ TEST CASE – APP VINAGO (MODULE SẢN PHẨM) (Cập nhật Server & Thiết bị)"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 30

    headers = [
        "Mã TC", "Chức năng", "Test trên", "Server", "Thiết bị", "Tiêu đề", "Loại kiểm thử",
        "Điều kiện", "Các bước", "Kết quả mong đợi", "Ưu tiên"
    ]
    col_widths = [14, 25, 10, 10, 12, 42, 16, 38, 50, 45, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    groups = [
        ("NHÓM 1: CẤU TRÚC DANH MỤC & SẢN PHẨM", "TC_SP_001", "TC_SP_011"),
        ("NHÓM 2: ẨN/HIỆN DANH MỤC SẢN PHẨM", "TC_SP_012", "TC_SP_016"),
        ("NHÓM 3: ẨN/HIỆN THEO ĐỐI TƯỢNG", "TC_SP_017", "TC_SP_025"),
        ("NHÓM 4: SẢN PHẨM BÁN CHẠY & MỚI", "TC_SP_026", "TC_SP_030"),
        ("NHÓM 5: SP TƯƠNG QUAN & KHÔNG HOA HỒNG", "TC_SP_031", "TC_SP_034"),
        ("NHÓM 6: GIAO DIỆN APP", "TC_SP_035", "TC_SP_040"),
    ]

    type_color = {
        "Luồng chuẩn":    COLOR_HAPPY,
        "Luồng ngoại lệ": COLOR_NEGATIVE,
        "Giá trị biên":   COLOR_BOUNDARY,
        "Logic hệ thống": COLOR_LOGIC,
    }
    priority_color = {"Cao": COLOR_HIGH, "Trung bình": COLOR_MEDIUM, "Thấp": COLOR_LOW}

    row_cursor = 3
    tc_ids = [tc["id"] for tc in test_cases]

    for group_name, start_id, end_id in groups:
        ws.merge_cells(f"A{row_cursor}:K{row_cursor}")
        g_cell = ws.cell(row=row_cursor, column=1, value=f"  {group_name}")
        g_cell.font, g_cell.fill, g_cell.alignment, g_cell.border = font_group, PatternFill("solid", fgColor=COLOR_GROUP_ROW), align_left_c, border_thick
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        start_idx = tc_ids.index(start_id)
        end_idx = tc_ids.index(end_id) + 1
        
        for tc in test_cases[start_idx:end_idx]:
            bg = type_color.get(tc["type"], "FFFFFF")
            values = [tc["id"], tc["module"], tc["platform"], tc["server"], tc["device"], tc["title"], tc["type"], tc["preconditions"], tc["steps"], tc["expected"], tc["priority"]]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.fill, cell.border = PatternFill("solid", fgColor=bg), border_all
                if col_idx == 1: cell.font, cell.alignment = font_id, align_center
                elif col_idx == 3: 
                    plt = tc["platform"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("274E13" if plt == "App" else "B45F06"))
                    cell.fill = PatternFill("solid", fgColor=(COLOR_APP if plt == "App" else COLOR_CMS))
                    cell.alignment = align_center
                elif col_idx in (4, 5):
                    cell.font = font_normal
                    cell.alignment = align_center
                elif col_idx == 7:
                    c_txt = tc["type"]
                    cell.font = Font(name="Calibri", bold=True, size=9, color=("276221" if c_txt=="Luồng chuẩn" else "7B2C00" if c_txt=="Luồng ngoại lệ" else "4B0082" if c_txt=="Logic hệ thống" else "7B6300"))
                    cell.alignment = align_center
                elif col_idx == 11:
                    cell.font, cell.fill, cell.alignment = Font(name="Calibri", bold=True, color="FFFFFF", size=9), PatternFill("solid", fgColor=priority_color.get(tc.get("priority", "Trung bình"), "000000")), align_center
                elif col_idx in (2, 6): cell.font, cell.alignment = font_normal, align_left_c
                else: cell.font, cell.alignment = font_normal, align_left
            ws.row_dimensions[row_cursor].height = 60
            row_cursor += 1
        row_cursor += 1 

    output_path = r"d:\Java lean\TestCase\TestCase_VINAGO_SanPham_TiengViet_Full.xlsx"
    wb.save(output_path)
    print("TC_SP_FULL_UPDATE_OK")

if __name__ == "__main__":
    create_excel()
