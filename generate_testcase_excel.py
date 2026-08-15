# -*- coding: utf-8 -*-
"""
Tạo file Excel Test Case cho App VINAGO – Module Khách Hàng
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── DỮ LIỆU TEST CASE ────────────────────────────────────────────────────────

test_cases = [
    # ── NHÓM 1: ĐĂNG KÝ TÀI KHOẢN (APP) ──────────────────────────────────────
    {
        "id": "TC001", "module": "Đăng ký TK – App",
        "title": "Đăng ký thành công với đầy đủ thông tin hợp lệ",
        "type": "Happy Path",
        "preconditions": "- App đang ở màn hình Đăng ký\n- SĐT 0901234567 chưa tồn tại\n- MST 0123456789 chưa tồn tại trong DS tài khoản",
        "steps": "1. Nhập Họ và tên: Nguyễn Văn A\n2. Nhập SĐT: 0901234567\n3. Nhập Tên cửa hàng: Cửa hàng Minh An\n4. Chọn Tỉnh/thành, Phường/xã\n5. Nhập Địa chỉ: 123 Nguyễn Huệ\n6. Nhập MST: 0123456789\n7. Chờ auto-fill Tên đơn vị, Địa chỉ từ Cổng Thuế\n8. Nhập Email (tùy chọn)\n9. Nhập Mật khẩu: 123456\n10. Nhập lại Mật khẩu: 123456\n11. Upload 1 ảnh hợp lệ\n12. Chọn ít nhất 1 danh mục SP\n13. Trả lời bộ câu hỏi khảo sát\n14. Nhấn Đăng ký",
        "expected": "- Đăng ký thành công\n- TK xuất hiện CMS với trạng thái Chờ duyệt\n- App hiển thị thông báo chờ duyệt\n- Chưa có Mã KH",
        "priority": "High"
    },
    {
        "id": "TC002", "module": "Đăng ký TK – App",
        "title": "Thông tin XHĐ tự động lấy từ Cổng Thuế khi nhập MST hợp lệ",
        "type": "Happy Path",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Điền các trường cơ bản hợp lệ\n2. Nhập MST có tồn tại trên Cổng Thuế: 0300000001\n3. Bấm ra ngoài ô MST",
        "expected": "- Tên đơn vị và Địa chỉ XHĐ tự động điền từ Cổng Thuế\n- Địa chỉ rã thành 3 cấp\n- Người dùng vẫn có thể chỉnh sửa thông tin auto-fill",
        "priority": "High"
    },
    {
        "id": "TC003", "module": "Đăng ký TK – App",
        "title": "Cổng Thuế không trả về dữ liệu – cho phép nhập thủ công Tên đơn vị in hoa",
        "type": "Happy Path",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Nhập MST không tồn tại trên Cổng Thuế: 9999999999\n2. Quan sát trường Tên đơn vị\n3. Nhập thủ công: ten don vi thu cong",
        "expected": "- Không crash ứng dụng\n- Cho phép nhập thủ công Tên đơn vị\n- Tên đơn vị mặc định IN HOA khi nhập\n- Vẫn tiếp tục đăng ký được",
        "priority": "Medium"
    },
    {
        "id": "TC004", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại khi bỏ trống trường Họ và tên (bắt buộc)",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Để trống trường Họ và tên\n2. Điền đầy đủ các trường còn lại hợp lệ\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Vui lòng nhập Họ và tên\n- Không submit được form",
        "priority": "High"
    },
    {
        "id": "TC005", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại khi bỏ trống Số điện thoại",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Để trống trường Số điện thoại\n2. Điền đầy đủ các trường còn lại\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Vui lòng nhập Số điện thoại\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC006", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại khi không chọn Tỉnh/thành hoặc Phường/xã",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Để trống Tỉnh/thành hoặc Phường/xã\n2. Điền đầy đủ các trường còn lại\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi yêu cầu chọn địa chỉ\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC007", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại khi bỏ trống Mã số thuế",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Để trống MST\n2. Điền đầy đủ các trường còn lại\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Vui lòng nhập Mã số thuế\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC008", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại: SĐT đã tồn tại trong DS tài khoản (Đang hoạt động)",
        "type": "Negative",
        "preconditions": "SĐT 0909090909 đã được đăng ký bởi TK đang hoạt động",
        "steps": "1. Nhập SĐT: 0909090909\n2. Điền đầy đủ các trường còn lại hợp lệ\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Số điện thoại đã được sử dụng\n- Không tạo được TK mới",
        "priority": "High"
    },
    {
        "id": "TC009", "module": "Đăng ký TK – App",
        "title": "Đăng ký thất bại: MST đã tồn tại trong DS tài khoản (Đang hoạt động)",
        "type": "Negative",
        "preconditions": "MST 0100111222 đã tồn tại trong DS tài khoản với trạng thái Đang hoạt động",
        "steps": "1. Nhập MST: 0100111222\n2. Điền đầy đủ các trường còn lại hợp lệ\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Mã số thuế đã được đăng ký\n- Không tạo được TK mới",
        "priority": "High"
    },
    {
        "id": "TC010", "module": "Đăng ký TK – App",
        "title": "Đăng ký thành công: MST trùng với TK đã bị Từ chối (không check trùng)",
        "type": "Happy Path",
        "preconditions": "MST 0200333444 tồn tại trong DS TK với trạng thái Từ chối",
        "steps": "1. Nhập MST: 0200333444\n2. Điền đầy đủ các trường còn lại hợp lệ\n3. Nhấn Đăng ký",
        "expected": "- Hệ thống cho phép đăng ký thành công (không báo trùng MST)\n- TK mới tạo với trạng thái Chờ duyệt",
        "priority": "Medium"
    },
    {
        "id": "TC011", "module": "Đăng ký TK – App",
        "title": "[Boundary] Mật khẩu 5 ký tự (Min-1) → thất bại",
        "type": "Boundary",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Nhập Mật khẩu: 12345 (5 ký tự)\n2. Nhập lại Mật khẩu: 12345\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Mật khẩu tối thiểu 6 ký tự\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC012", "module": "Đăng ký TK – App",
        "title": "[Boundary] Mật khẩu 6 ký tự (Min) → thành công",
        "type": "Boundary",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Nhập Mật khẩu: 123456 (6 ký tự)\n2. Nhập lại Mật khẩu: 123456\n3. Điền đầy đủ các trường còn lại hợp lệ\n4. Nhấn Đăng ký",
        "expected": "- Hệ thống chấp nhận mật khẩu 6 ký tự\n- Đăng ký thành công",
        "priority": "High"
    },
    {
        "id": "TC013", "module": "Đăng ký TK – App",
        "title": "Mật khẩu và Nhập lại mật khẩu không khớp → thất bại",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Nhập Mật khẩu: 123456\n2. Nhập lại Mật khẩu: 654321\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Mật khẩu nhập lại không khớp\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC014", "module": "Đăng ký TK – App",
        "title": "[Boundary] Upload 0 hình ảnh → thất bại",
        "type": "Boundary",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Điền đầy đủ các trường hợp lệ\n2. Không upload bất kỳ hình ảnh nào\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Vui lòng thêm ít nhất 1 hình ảnh\n- Không submit được",
        "priority": "Medium"
    },
    {
        "id": "TC015", "module": "Đăng ký TK – App",
        "title": "[Boundary] Upload 5 hình ảnh (Max) → thành công, hình đầu làm avatar",
        "type": "Boundary",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Upload đúng 5 ảnh hợp lệ\n2. Điền đầy đủ các trường còn lại\n3. Nhấn Đăng ký",
        "expected": "- Upload thành công\n- Hình đầu tiên được đặt làm avatar mặc định",
        "priority": "Medium"
    },
    {
        "id": "TC016", "module": "Đăng ký TK – App",
        "title": "[Boundary] Upload 6 hình ảnh (Max+1) → không cho phép",
        "type": "Boundary",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Thử upload 6 ảnh hợp lệ",
        "expected": "- Hệ thống không cho upload quá 5 hình\n- Hiển thị thông báo giới hạn tối đa 5 hình",
        "priority": "Medium"
    },
    {
        "id": "TC017", "module": "Đăng ký TK – App",
        "title": "Upload file không phải ảnh (PDF) → thất bại",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Chọn file document.pdf để upload ảnh",
        "expected": "- Hiển thị lỗi: Định dạng file không hợp lệ\n- Không cho phép upload",
        "priority": "Medium"
    },
    {
        "id": "TC018", "module": "Đăng ký TK – App",
        "title": "Không chọn danh mục sản phẩm → thất bại",
        "type": "Negative",
        "preconditions": "App đang ở màn hình Đăng ký",
        "steps": "1. Điền đầy đủ tất cả trường bắt buộc hợp lệ\n2. Không chọn bất kỳ danh mục sản phẩm nào\n3. Nhấn Đăng ký",
        "expected": "- Hiển thị lỗi: Vui lòng chọn ít nhất 1 danh mục sản phẩm\n- Không submit được",
        "priority": "High"
    },
    {
        "id": "TC019", "module": "Đăng ký TK – App",
        "title": "Chọn 3 danh mục SP → hiển thị đúng 3 lần bộ câu hỏi khảo sát",
        "type": "Happy Path",
        "preconditions": "App đang ở màn hình Đăng ký, hệ thống có ≥3 danh mục cấp 1",
        "steps": "1. Điền đầy đủ thông tin hợp lệ\n2. Chọn 3 danh mục sản phẩm (A, B, C)\n3. Quan sát bộ câu hỏi khảo sát",
        "expected": "- Hệ thống hiển thị 3 lần bộ câu hỏi, mỗi lần đúng 1 danh mục\n- Phải trả lời đủ 3 bộ mới tiếp tục được",
        "priority": "Medium"
    },

    # ── NHÓM 2: QUẢN LÝ TRẠNG THÁI TÀI KHOẢN (CMS) ───────────────────────────
    {
        "id": "TC020", "module": "Quản lý TK – CMS",
        "title": "Duyệt TK Chờ duyệt → Đang hoạt động (MST chưa có trong DS người mua)",
        "type": "Happy Path",
        "preconditions": "- Đăng nhập CMS quyền Admin\n- TK KH trạng thái Chờ duyệt\n- MST chưa tồn tại trong DS người mua",
        "steps": "1. CMS → DS Tài khoản\n2. Tìm KH trạng thái Chờ duyệt\n3. Nhấn nút Duyệt\n4. Xác nhận duyệt",
        "expected": "- Trạng thái TK → Đang hoạt động\n- Hệ thống tự sinh Mã KH và Nickname\n- Loại KH mặc định = Khách bán Lẻ\n- Mặc định khóa giá bán, CS/CT\n- Hệ thống tự tạo mới thông tin Người mua\n- App KH: hiển thị đầy đủ tất cả Tab",
        "priority": "High"
    },
    {
        "id": "TC021", "module": "Quản lý TK – CMS",
        "title": "Duyệt TK → MST đã có trong DS người mua → cập nhật (không tạo mới)",
        "type": "Happy Path",
        "preconditions": "- TK KH trạng thái Chờ duyệt với MST 0123456789\n- MST 0123456789 đã có trong DS người mua",
        "steps": "1. CMS → DS Tài khoản\n2. Nhấn Duyệt TK KH\n3. Xác nhận",
        "expected": "- Hệ thống CẬP NHẬT (không tạo mới) bản ghi người mua\n- Cập nhật: Mã KH, Tên KH, Loại KH, Trạng thái TK\n- KHÔNG thay đổi: Trạng thái người mua, Tên đơn vị, Địa chỉ, Email",
        "priority": "High"
    },
    {
        "id": "TC022", "module": "Quản lý TK – CMS",
        "title": "Từ chối TK Chờ duyệt → trạng thái Từ chối",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK KH trạng thái Chờ duyệt",
        "steps": "1. Tìm KH trạng thái Chờ duyệt\n2. Nhấn nút Từ chối\n3. Nhập lý do và xác nhận",
        "expected": "- Trạng thái TK → Từ chối\n- Không sinh ra Mã KH\n- TK bị Từ chối không check trùng SĐT/MST với đăng ký mới",
        "priority": "Medium"
    },
    {
        "id": "TC023", "module": "Quản lý TK – CMS",
        "title": "Tạm dừng TK Đang hoạt động → Ngưng hoạt động",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK KH trạng thái Đang hoạt động",
        "steps": "1. Tìm KH trạng thái Đang hoạt động\n2. Nhấn nút Tạm dừng\n3. Xác nhận",
        "expected": "- Trạng thái TK → Ngưng hoạt động\n- KH không đăng nhập App được\n- KH Sỉ: không đặt hàng được\n- KH Lẻ: không mua trực tiếp CT, có thể mua qua KH Sỉ",
        "priority": "High"
    },
    {
        "id": "TC024", "module": "Quản lý TK – CMS",
        "title": "Kích hoạt TK Ngưng hoạt động → Đang hoạt động",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK KH trạng thái Ngưng hoạt động",
        "steps": "1. Tìm KH trạng thái Ngưng hoạt động\n2. Nhấn nút Kích hoạt\n3. Xác nhận",
        "expected": "- Trạng thái TK → Đang hoạt động\n- KH đăng nhập và đặt hàng bình thường trên App",
        "priority": "Medium"
    },
    {
        "id": "TC025", "module": "Quản lý TK – CMS",
        "title": "Nút chức năng hiển thị đúng theo từng trạng thái TK",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK ở 4 trạng thái: Chờ duyệt, Đang HĐ, Ngưng HĐ, Từ chối",
        "steps": "1. Mở từng TK theo từng trạng thái\n2. Quan sát danh sách nút chức năng",
        "expected": "- Chờ duyệt: Duyệt, Từ chối, Điều chỉnh\n- Đang hoạt động: Tạm dừng, Điều chỉnh\n- Ngưng hoạt động: Kích hoạt\n- Từ chối: không có nút",
        "priority": "High"
    },
    {
        "id": "TC026", "module": "Quản lý TK – CMS",
        "title": "Tạo mới TK từ CMS với Loại KH = Khách bán Sỉ",
        "type": "Happy Path",
        "preconditions": "- Đăng nhập CMS quyền Admin\n- MST 0500123456 chưa tồn tại",
        "steps": "1. CMS → DS Tài khoản → Tạo mới\n2. Chọn Loại KH: Khách bán Sỉ\n3. Nhập MST: 0500123456\n4. Chờ auto-fill thông tin XHĐ từ Cổng Thuế\n5. Điền đầy đủ các trường còn lại\n6. Nhấn Lưu",
        "expected": "- TK được tạo thành công\n- Loại KH = Khách bán Sỉ\n- Thông tin XHĐ được lấy từ Cổng Thuế",
        "priority": "Medium"
    },
    {
        "id": "TC027", "module": "Quản lý TK – CMS",
        "title": "Điều chỉnh thông tin TK Đang hoạt động",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK KH trạng thái Đang hoạt động",
        "steps": "1. Tìm KH Đang hoạt động\n2. Nhấn Điều chỉnh\n3. Sửa trường Tên cửa hàng\n4. Nhấn Lưu",
        "expected": "- Thông tin được cập nhật thành công\n- Lịch sử thay đổi được ghi lại",
        "priority": "Medium"
    },
    {
        "id": "TC028", "module": "Quản lý TK – CMS",
        "title": "[Boundary] Nickname KH: 15 ký tự (Max) → thành công",
        "type": "Boundary",
        "preconditions": "TK KH đang ở trạng thái Đang hoạt động",
        "steps": "1. Vào chi tiết KH\n2. Nhập Nickname: ABCDEFGHIJKLMNO (15 ký tự)\n3. Lưu",
        "expected": "- Lưu thành công",
        "priority": "Medium"
    },
    {
        "id": "TC029", "module": "Quản lý TK – CMS",
        "title": "[Boundary] Nickname KH: 16 ký tự (Max+1) → thất bại hoặc tự cắt",
        "type": "Boundary",
        "preconditions": "TK KH đang ở trạng thái Đang hoạt động",
        "steps": "1. Vào chi tiết KH\n2. Nhập Nickname: ABCDEFGHIJKLMNOP (16 ký tự)",
        "expected": "- Hệ thống không cho nhập quá 15 ký tự\n- Hoặc tự cắt tại ký tự thứ 15",
        "priority": "Medium"
    },
    {
        "id": "TC030", "module": "Quản lý TK – CMS",
        "title": "Nickname nhập ký tự đặc biệt từ App → bị chặn",
        "type": "Negative",
        "preconditions": "TK KH đang đăng nhập App",
        "steps": "1. Vào mục Nickname trên App\n2. Nhập ký tự đặc biệt: Nick@#$",
        "expected": "- App tự xử lý: không cho nhập ký tự đặc biệt\n- Chỉ cho phép ký tự chữ và số",
        "priority": "Medium"
    },
    {
        "id": "TC031", "module": "Quản lý TK – CMS",
        "title": "Xuất dữ liệu DS tài khoản theo bộ lọc",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS quyền Admin",
        "steps": "1. Lọc DS theo Trạng thái = Đang hoạt động\n2. Nhấn Xuất dữ liệu",
        "expected": "- File export (Excel/CSV) tải về thành công\n- Dữ liệu trong file khớp với dữ liệu đang lọc trên màn hình",
        "priority": "Low"
    },

    # ── NHÓM 3: TÌM KIẾM & BỘ LỌC ─────────────────────────────────────────────
    {
        "id": "TC032", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Tìm kiếm theo Mã KH – kết quả chính xác",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại KH có Mã KH = KH0001",
        "steps": "1. Vào DS Tài khoản\n2. Nhập vào thanh tìm kiếm: KH0001\n3. Nhấn Enter",
        "expected": "- DS hiển thị đúng KH có Mã KH = KH0001\n- Nếu không tìm thấy → hiển thị thông báo không có kết quả",
        "priority": "Medium"
    },
    {
        "id": "TC033", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Tìm kiếm theo Tên KH – partial match, không phân biệt hoa thường",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại nhiều KH có tên chứa từ Minh",
        "steps": "1. Nhập từ khóa: minh (chữ thường)\n2. Nhấn Enter",
        "expected": "- Hiển thị tất cả KH có tên chứa Minh (không phân biệt hoa/thường)",
        "priority": "Medium"
    },
    {
        "id": "TC034", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Tìm kiếm theo SĐT đăng nhập",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại KH có SĐT = 0901234567",
        "steps": "1. Nhập SĐT: 0901234567 vào thanh tìm kiếm",
        "expected": "- DS hiển thị đúng KH có SĐT = 0901234567",
        "priority": "Medium"
    },
    {
        "id": "TC035", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Tìm kiếm theo MST",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại KH có MST = 0123456789",
        "steps": "1. Nhập MST: 0123456789 vào thanh tìm kiếm",
        "expected": "- DS hiển thị đúng KH có MST = 0123456789",
        "priority": "Medium"
    },
    {
        "id": "TC036", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Tìm kiếm không có kết quả → hiển thị thông báo",
        "type": "Negative",
        "preconditions": "Đăng nhập CMS",
        "steps": "1. Nhập từ khóa không tồn tại: XXXXXXXX",
        "expected": "- DS trống\n- Hiển thị thông báo: Không tìm thấy kết quả phù hợp",
        "priority": "Low"
    },
    {
        "id": "TC037", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Bộ lọc Loại KH = Khách bán Sỉ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. DS có cả KH Sỉ và KH Lẻ",
        "steps": "1. Chọn bộ lọc Loại KH = Khách bán Sỉ",
        "expected": "- DS chỉ hiển thị KH Sỉ",
        "priority": "Medium"
    },
    {
        "id": "TC038", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Bộ lọc Loại KH = Khách bán Lẻ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS",
        "steps": "1. Chọn bộ lọc Loại KH = Khách bán Lẻ",
        "expected": "- DS chỉ hiển thị KH Lẻ",
        "priority": "Medium"
    },
    {
        "id": "TC039", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Bộ lọc Trạng thái TK – kiểm tra lần lượt 4 trạng thái",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. DS có TK ở đủ 4 trạng thái",
        "steps": "1. Lọc lần lượt: Chờ duyệt / Đang hoạt động / Ngưng hoạt động / Từ chối",
        "expected": "- Mỗi lần lọc chỉ hiển thị TK đúng trạng thái được chọn",
        "priority": "Medium"
    },
    {
        "id": "TC040", "module": "Tìm kiếm & Lọc – DS TK",
        "title": "Kết hợp bộ lọc: Loại KH = Sỉ + Trạng thái = Đang hoạt động",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS",
        "steps": "1. Chọn Loại KH = Khách bán Sỉ\n2. Chọn Trạng thái = Đang hoạt động\n3. Nhấn tìm kiếm",
        "expected": "- Chỉ hiển thị KH vừa là Sỉ vừa đang Hoạt động",
        "priority": "Medium"
    },

    # ── NHÓM 4: DANH SÁCH NGƯỜI MUA ────────────────────────────────────────────
    {
        "id": "TC041", "module": "DS Người mua – CMS",
        "title": "Tạo mới Người mua thành công từ CMS",
        "type": "Happy Path",
        "preconditions": "- Đăng nhập CMS quyền Admin\n- MST 0700123456 chưa tồn tại trong DS người mua",
        "steps": "1. CMS → DS Người mua → Tạo mới\n2. Nhập MST: 0700123456\n3. Nhập Tên đơn vị: CONG TY TNHH TEST\n4. Trường Tên KH tự điền theo Tên đơn vị\n5. Chọn Tỉnh/Thành, Phường/Xã, Địa chỉ\n6. Nhập Email\n7. Nhấn Lưu",
        "expected": "- Người mua tạo thành công với trạng thái Nháp\n- Tên KH mặc định = Tên đơn vị\n- Tên đơn vị lưu ở dạng IN HOA",
        "priority": "High"
    },
    {
        "id": "TC042", "module": "DS Người mua – CMS",
        "title": "Tạo mới Người mua thất bại: MST đã tồn tại",
        "type": "Negative",
        "preconditions": "MST 0123456789 đã tồn tại trong DS người mua (trạng thái Đang giao dịch)",
        "steps": "1. CMS → DS Người mua → Tạo mới\n2. Nhập MST: 0123456789\n3. Nhấn Lưu",
        "expected": "- Hiển thị lỗi: Mã số thuế đã tồn tại\n- Không tạo được bản ghi mới",
        "priority": "High"
    },
    {
        "id": "TC043", "module": "DS Người mua – CMS",
        "title": "Luồng trạng thái Người mua: Nháp → Chờ xác nhận → Đang giao dịch",
        "type": "Happy Path",
        "preconditions": "Tồn tại Người mua trạng thái Nháp",
        "steps": "1. Tìm Người mua trạng thái Nháp\n2. Nhấn Xác nhận → trạng thái chuyển sang Chờ xác nhận\n3. Nhấn Duyệt → hệ thống truyền Bravo → trạng thái Đang giao dịch",
        "expected": "- Luồng đúng: Nháp → Chờ xác nhận → Đang giao dịch\n- Tại Nháp: nút Điều chỉnh, Xác nhận\n- Tại Chờ XN: nút Duyệt, Từ chối, Điều chỉnh\n- Tại Đang GD: nút Tạm dừng, Điều chỉnh",
        "priority": "High"
    },
    {
        "id": "TC044", "module": "DS Người mua – CMS",
        "title": "Nút chức năng Người mua đúng theo từng trạng thái (5 trạng thái)",
        "type": "Happy Path",
        "preconditions": "Tồn tại người mua ở cả 5 trạng thái: Nháp, Chờ XN, Đang GD, Ngưng GD, Hủy",
        "steps": "1. Mở lần lượt từng người mua ở 5 trạng thái\n2. Quan sát nút chức năng",
        "expected": "- Nháp: Điều chỉnh, Xác nhận\n- Chờ xác nhận: Duyệt, Từ chối, Điều chỉnh\n- Đang giao dịch: Tạm dừng, Điều chỉnh\n- Ngưng giao dịch: Kích hoạt\n- Hủy: không có nút",
        "priority": "High"
    },
    {
        "id": "TC045", "module": "DS Người mua – CMS",
        "title": "Tạm dừng Người mua Đang giao dịch → Ngưng giao dịch",
        "type": "Happy Path",
        "preconditions": "Tồn tại Người mua trạng thái Đang giao dịch",
        "steps": "1. Tìm Người mua trạng thái Đang giao dịch\n2. Nhấn Tạm dừng\n3. Xác nhận",
        "expected": "- Trạng thái → Ngưng giao dịch\n- Đơn hàng mới với MST này hiển thị cảnh báo không giao dịch",
        "priority": "Medium"
    },
    {
        "id": "TC046", "module": "DS Người mua – CMS",
        "title": "Kích hoạt Người mua Ngưng giao dịch → Đang giao dịch",
        "type": "Happy Path",
        "preconditions": "Tồn tại Người mua trạng thái Ngưng giao dịch",
        "steps": "1. Tìm Người mua trạng thái Ngưng giao dịch\n2. Nhấn Kích hoạt",
        "expected": "- Trạng thái → Đang giao dịch",
        "priority": "Medium"
    },
    {
        "id": "TC047", "module": "DS Người mua – CMS",
        "title": "Không được thay đổi Loại KH trong DS Người mua",
        "type": "Negative",
        "preconditions": "Đăng nhập CMS. Tồn tại Người mua Đang giao dịch",
        "steps": "1. Vào chi tiết Người mua Đang giao dịch\n2. Thử thay đổi trường Loại KH",
        "expected": "- Trường Loại KH ở trạng thái read-only hoặc không có tùy chọn thay đổi",
        "priority": "Medium"
    },
    {
        "id": "TC048", "module": "DS Người mua – CMS",
        "title": "Tự động tạo Người mua khi duyệt TK App (MST chưa có trong DS)",
        "type": "Happy Path",
        "preconditions": "- TK KH trạng thái Chờ duyệt với MST 0900123456\n- MST 0900123456 chưa có trong DS người mua",
        "steps": "1. Admin duyệt TK KH",
        "expected": "- DS Người mua tự động có bản ghi mới với MST 0900123456\n- Thông tin theo XHĐ của TK\n- Trạng thái: Chờ xác nhận → Bravo → Đang giao dịch",
        "priority": "High"
    },
    {
        "id": "TC049", "module": "DS Người mua – CMS",
        "title": "Tìm kiếm Người mua theo Tên XHĐ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại Người mua có Tên XHĐ = Công ty ABC",
        "steps": "1. DS Người mua → Thanh tìm kiếm\n2. Nhập: Công ty ABC",
        "expected": "- Hiển thị đúng Người mua có Tên XHĐ = Công ty ABC",
        "priority": "Medium"
    },
    {
        "id": "TC050", "module": "DS Người mua – CMS",
        "title": "Bộ lọc Trạng thái Người mua – kiểm tra 4 trạng thái",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. DS có đủ 4 trạng thái",
        "steps": "1. Lọc lần lượt: Chờ xác nhận / Đang giao dịch / Ngưng giao dịch / Hủy",
        "expected": "- Mỗi lần lọc chỉ hiển thị người mua đúng trạng thái đã chọn",
        "priority": "Medium"
    },

    # ── NHÓM 5: TÀI KHOẢN CHÍNH/PHỤ ──────────────────────────────────────────
    {
        "id": "TC051", "module": "TK Chính/Phụ – App",
        "title": "TK chính tạo TK phụ thành công từ App",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK chính đang hoạt động",
        "steps": "1. Vào Quản lý TK → Tạo TK phụ\n2. Nhập Tên TK: Nhân Viên Bán Hàng 1\n3. Nhập SĐT: 0912345678\n4. Phân quyền cho TK phụ\n5. Nhấn Tạo",
        "expected": "- TK phụ được tạo trạng thái Chờ kích hoạt\n- OTP gửi về SĐT của TK CHÍNH (không phải SĐT phụ)\n- CMS hiển thị TK phụ để Admin kích hoạt/từ chối",
        "priority": "High"
    },
    {
        "id": "TC052", "module": "TK Chính/Phụ – App",
        "title": "TK chính tạo không giới hạn số lượng TK phụ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK chính đang hoạt động",
        "steps": "1. Tạo liên tiếp 10 TK phụ với SĐT khác nhau",
        "expected": "- Hệ thống không báo giới hạn\n- Tất cả TK phụ được tạo thành công (Chờ kích hoạt)",
        "priority": "Medium"
    },
    {
        "id": "TC053", "module": "TK Chính/Phụ – CMS",
        "title": "Admin CMS kích hoạt TK phụ Chờ kích hoạt → Đang hoạt động",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK phụ trạng thái Chờ kích hoạt",
        "steps": "1. CMS → Chi tiết KH → Tab Quản lý TK → TK phụ\n2. Nhấn Kích hoạt TK phụ",
        "expected": "- TK phụ chuyển sang trạng thái Đang hoạt động",
        "priority": "High"
    },
    {
        "id": "TC054", "module": "TK Chính/Phụ – CMS",
        "title": "Admin CMS từ chối TK phụ → TK chính tạo lại được với cùng SĐT",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS. Tồn tại TK phụ trạng thái Chờ kích hoạt",
        "steps": "1. CMS → TK phụ Chờ kích hoạt → Nhấn Từ chối\n2. TK chính thử tạo lại TK phụ với cùng SĐT đã bị từ chối",
        "expected": "- TK phụ bị từ chối\n- TK chính có thể tạo lại TK phụ với cùng SĐT (không check trùng SĐT bị từ chối)",
        "priority": "Medium"
    },
    {
        "id": "TC055", "module": "TK Chính/Phụ – App",
        "title": "TK phụ có quyền Không thấy công nợ → mục Công nợ bị ẩn",
        "type": "Negative",
        "preconditions": "TK phụ được phân quyền Không thấy thông tin công nợ",
        "steps": "1. Đăng nhập App bằng TK phụ\n2. Truy cập mục Công nợ",
        "expected": "- Mục Công nợ không hiển thị hoặc bị ẩn/mờ",
        "priority": "High"
    },
    {
        "id": "TC056", "module": "TK Chính/Phụ – App",
        "title": "TK phụ có quyền Không cho phép Đặt hàng → ẩn xe hàng và nút thêm",
        "type": "Negative",
        "preconditions": "TK phụ được phân quyền Không cho phép Đặt hàng",
        "steps": "1. Đăng nhập App bằng TK phụ\n2. Truy cập trang Sản phẩm",
        "expected": "- Nút Thêm vào xe hàng bị mờ/ẩn\n- Icon thêm nhanh không hiển thị\n- Tab Xe hàng không thấy",
        "priority": "High"
    },
    {
        "id": "TC057", "module": "TK Chính/Phụ – App",
        "title": "TK phụ mới có đầy đủ quyền mặc định",
        "type": "Happy Path",
        "preconditions": "TK phụ mới được tạo, không điều chỉnh phân quyền",
        "steps": "1. Đăng nhập App bằng TK phụ mới\n2. Truy cập các mục: Công nợ, Giá bán, CS/CT, Xe hàng",
        "expected": "- TK phụ có đầy đủ quyền mặc định (thấy tất cả mục)",
        "priority": "Medium"
    },
    {
        "id": "TC058", "module": "TK Chính/Phụ – App",
        "title": "TK phụ nhập chênh lệch giá bán và kỳ hạn nợ khi đặt hàng cho tuyến dưới",
        "type": "Happy Path",
        "preconditions": "TK phụ đang hoạt động, không bị giới hạn quyền đặt hàng",
        "steps": "1. TK phụ đặt hàng cho tuyến dưới\n2. Nhập giá bán khác giá mặc định\n3. Nhập kỳ hạn nợ",
        "expected": "- Hệ thống cho phép nhập chênh lệch giá bán và kỳ hạn nợ",
        "priority": "Medium"
    },
    {
        "id": "TC059", "module": "TK Chính/Phụ – CMS",
        "title": "Điều chỉnh quyền TK phụ Đang hoạt động – áp dụng ngay trên App",
        "type": "Happy Path",
        "preconditions": "TK phụ trạng thái Đang hoạt động",
        "steps": "1. CMS → Chi tiết KH → TK phụ → Chỉnh sửa quyền\n2. Bật quyền Không thấy giá bán\n3. Lưu",
        "expected": "- Thay đổi quyền lưu thành công\n- TK phụ ngay lập tức không thấy giá bán trên App",
        "priority": "Medium"
    },
    {
        "id": "TC060", "module": "TK Chính/Phụ – App",
        "title": "Quên mật khẩu TK phụ → OTP gửi về SĐT TK chính",
        "type": "Happy Path",
        "preconditions": "TK phụ đang hoạt động",
        "steps": "1. App: Đăng nhập bằng SĐT phụ\n2. Nhấn Quên mật khẩu\n3. Quan sát OTP gửi về SĐT nào",
        "expected": "- OTP gửi về SĐT của TK CHÍNH (không phải SĐT phụ)",
        "priority": "High"
    },

    # ── NHÓM 6: CẤP BẬC ────────────────────────────────────────────────────────
    {
        "id": "TC061", "module": "Cấp bậc – KH Sỉ",
        "title": "Cấp bậc mặc định khi TK Sỉ được duyệt là Thạch Anh",
        "type": "Happy Path",
        "preconditions": "KH được phân loại là Khách bán Sỉ và vừa được Admin duyệt",
        "steps": "1. Admin duyệt TK KH với Loại KH = Khách bán Sỉ\n2. Kiểm tra cấp bậc trong CMS và App",
        "expected": "- Cấp bậc mặc định = Thạch Anh (cấp thấp nhất)\n- Dữ liệu xét cấp bậc tính từ đầu",
        "priority": "High"
    },
    {
        "id": "TC062", "module": "Cấp bậc – KH Lẻ",
        "title": "Khách bán Lẻ không có cấp bậc, chỉ hiển thị Thành viên",
        "type": "Negative",
        "preconditions": "Đăng nhập App bằng TK Khách bán Lẻ đang hoạt động",
        "steps": "1. Đăng nhập App bằng TK Khách bán Lẻ\n2. Truy cập mục Cấp bậc / Tài khoản",
        "expected": "- Hiển thị Thành viên (không có cấp bậc)\n- Không hiển thị các cấp Thạch Anh, Hổ Phách,...",
        "priority": "Medium"
    },
    {
        "id": "TC063", "module": "Cấp bậc – KH Sỉ",
        "title": "Hiển thị đúng cấp bậc hiện tại và cấp kế tiếp trên App",
        "type": "Happy Path",
        "preconditions": "KH Sỉ đang ở cấp Hổ Phách",
        "steps": "1. Đăng nhập App bằng TK KH Sỉ cấp Hổ Phách\n2. Vào mục Cấp bậc",
        "expected": "- Hiển thị cấp hiện tại = Hổ Phách + mức chiết khấu hiện tại\n- Hiển thị cấp kế tiếp = Ngọc Bích + mức chiết khấu tiếp theo\n- Hiển thị điều kiện đạt cấp Ngọc Bích (DTT, Tiền thu, SL người mua hợp lệ)",
        "priority": "Medium"
    },
    {
        "id": "TC064", "module": "Cấp bậc – KH Sỉ",
        "title": "Cấp bậc cập nhật vào 23:00 ngày cuối tháng, không reset dữ liệu cuối năm",
        "type": "Happy Path",
        "preconditions": "Hệ thống có job tự động cập nhật cấp bậc",
        "steps": "1. Kiểm tra thời gian job chạy cập nhật cấp bậc\n2. Kiểm tra dữ liệu cấp bậc sau khi sang năm mới",
        "expected": "- Job cập nhật chạy lúc 23:00 ngày cuối tháng\n- Cuối năm: dữ liệu cấp bậc KHÔNG bị reset",
        "priority": "High"
    },
    {
        "id": "TC065", "module": "Cấp bậc – KH Sỉ",
        "title": "5 cấp bậc KH Sỉ hiển thị đúng thứ tự từ thấp đến cao",
        "type": "Happy Path",
        "preconditions": "Đăng nhập CMS hoặc App bằng TK Sỉ",
        "steps": "1. Xem danh sách cấp bậc trên App hoặc CMS",
        "expected": "- Thứ tự đúng: Thạch Anh → Hổ Phách → Ngọc Bích → Hồng Ngọc → Thiên Thạch",
        "priority": "Medium"
    },
    {
        "id": "TC066", "module": "Cấp bậc – KH Sỉ",
        "title": "[Boundary] DTT đạt ngưỡng nhưng Tiền thu chưa đủ → không thăng cấp",
        "type": "Boundary",
        "preconditions": "KH Sỉ có DTT đạt ngưỡng nhưng Tiền thu < ngưỡng tối thiểu",
        "steps": "1. Kiểm tra dữ liệu KH Sỉ\n2. Chạy job cập nhật cấp bậc cuối tháng",
        "expected": "- KH KHÔNG được thăng cấp",
        "priority": "High"
    },
    {
        "id": "TC067", "module": "Cấp bậc – KH Sỉ",
        "title": "[Boundary] Cả DTT và Tiền thu đều đạt ngưỡng → thăng cấp",
        "type": "Boundary",
        "preconditions": "KH Sỉ có DTT ≥ ngưỡng và Tiền thu ≥ ngưỡng tối thiểu của cấp kế tiếp",
        "steps": "1. Kiểm tra dữ liệu KH Sỉ\n2. Chạy job cập nhật cấp bậc cuối tháng",
        "expected": "- KH được thăng lên cấp kế tiếp",
        "priority": "High"
    },
    {
        "id": "TC068", "module": "Cấp bậc – KH Sỉ",
        "title": "[Boundary] Cấp bậc reset về Thạch Anh khi chuyển Sỉ → Lẻ → Sỉ",
        "type": "Boundary",
        "preconditions": "KH Sỉ cấp Hổ Phách, sau đó chuyển sang Lẻ rồi quay lại Sỉ",
        "steps": "1. KH Sỉ cấp Hổ Phách\n2. Chuyển sang Lẻ\n3. Chuyển lại thành Sỉ\n4. Kiểm tra cấp bậc",
        "expected": "- Cấp bậc reset về Thạch Anh (mặc định)\n- Dữ liệu xét cấp bậc tính lại từ đầu (dữ liệu trước không tính)",
        "priority": "High"
    },

    # ── NHÓM 7: CHUYỂN ĐỔI LOẠI KH ────────────────────────────────────────────
    {
        "id": "TC069", "module": "Chuyển đổi Loại KH",
        "title": "Chuyển từ KH Lẻ → KH Sỉ: cấp bậc về Thạch Anh, ghi lịch sử",
        "type": "Happy Path",
        "preconditions": "- Đăng nhập CMS quyền Admin\n- KH đang là Khách bán Lẻ, trạng thái Đang hoạt động",
        "steps": "1. CMS → Chi tiết KH → Điều chỉnh Loại KH\n2. Chuyển sang Khách bán Sỉ\n3. Lưu",
        "expected": "- Loại KH = Khách bán Sỉ\n- Cấp bậc mặc định = Thạch Anh\n- Dữ liệu xét cấp bậc tính từ thời điểm chuyển\n- Lịch sử thay đổi phân loại được ghi lại",
        "priority": "High"
    },
    {
        "id": "TC070", "module": "Chuyển đổi Loại KH",
        "title": "Chuyển từ KH Sỉ → KH Lẻ: cấp bậc về Thành viên, phí bảo lãnh cấn trừ",
        "type": "Happy Path",
        "preconditions": "- Đăng nhập CMS quyền Admin\n- KH đang là Khách bán Sỉ có phí bảo lãnh và đơn hàng chưa TT",
        "steps": "1. CMS → Chi tiết KH Sỉ → Đổi Loại KH sang Khách bán Lẻ\n2. Lưu",
        "expected": "- Loại KH = Khách bán Lẻ\n- Cấp bậc = Thành viên\n- Phí bảo lãnh cấn trừ hết cho đơn hàng chưa TT\n- Phần dư chuyển về Công nợ\n- Lịch sử thay đổi được ghi lại",
        "priority": "High"
    },
    {
        "id": "TC071", "module": "Chuyển đổi Loại KH",
        "title": "[Boundary] Chuyển Sỉ → Lẻ khi không có phí bảo lãnh",
        "type": "Boundary",
        "preconditions": "KH Sỉ không có phí bảo lãnh",
        "steps": "1. Chuyển KH từ Sỉ sang Lẻ",
        "expected": "- Chuyển thành công\n- Không có cấn trừ phí bảo lãnh\n- Không phát sinh thêm công nợ",
        "priority": "Medium"
    },
    {
        "id": "TC072", "module": "Chuyển đổi Loại KH",
        "title": "[Boundary] Phí bảo lãnh dư khi Sỉ → Lẻ → chuyển đúng về công nợ",
        "type": "Boundary",
        "preconditions": "KH Sỉ: phí bảo lãnh = 10.000.000 VNĐ, đơn chưa TT = 8.000.000 VNĐ",
        "steps": "1. Chuyển KH từ Sỉ sang Lẻ\n2. Kiểm tra công nợ sau khi chuyển",
        "expected": "- Phí bảo lãnh cấn trừ 8.000.000 VNĐ cho đơn chưa TT\n- Phần dư = 2.000.000 VNĐ được chuyển về Công nợ",
        "priority": "High"
    },
    {
        "id": "TC073", "module": "Chuyển đổi Loại KH",
        "title": "[Boundary] XHĐ cũ vẫn lưu khi Sỉ → Lẻ nhưng không chọn được khi đặt hàng",
        "type": "Boundary",
        "preconditions": "KH Sỉ có 3 thông tin XHĐ (1 của KH + 2 người mua khác)",
        "steps": "1. Chuyển KH từ Sỉ sang Lẻ\n2. KH Lẻ thử đặt hàng và chọn thông tin XHĐ",
        "expected": "- KH Lẻ vẫn lưu các XHĐ cũ (không xóa)\n- Khi đặt hàng chỉ được chọn 1 XHĐ của chính KH Lẻ",
        "priority": "Medium"
    },
    {
        "id": "TC074", "module": "Chuyển đổi Loại KH",
        "title": "Chuyển Lẻ → Sỉ (lần 2): dữ liệu cấp bậc tính từ đầu, không cộng dồn",
        "type": "Happy Path",
        "preconditions": "KH đã từng là Sỉ, chuyển sang Lẻ, nay chuyển lại thành Sỉ",
        "steps": "1. Chuyển KH từ Lẻ lại thành Sỉ\n2. Kiểm tra cấp bậc và dữ liệu xét cấp bậc",
        "expected": "- Cấp bậc = Thạch Anh (reset)\n- Dữ liệu xét cấp bậc tính từ đầu (không cộng dồn dữ liệu trước)",
        "priority": "High"
    },
    {
        "id": "TC075", "module": "Chuyển đổi Loại KH",
        "title": "KH Sỉ: XHĐ của chính KH luôn là dòng đầu tiên trong tab Thông tin người mua",
        "type": "Happy Path",
        "preconditions": "KH Sỉ đã phát sinh đơn hàng với nhiều người mua khác nhau",
        "steps": "1. CMS → Chi tiết KH Sỉ → Tab Thông tin người mua",
        "expected": "- Dòng đầu tiên = XHĐ của chính KH Sỉ (có nhận diện riêng)\n- Các dòng tiếp theo = XHĐ người mua đã phát sinh đơn hàng và ghi nhận công nợ",
        "priority": "Medium"
    },

    # ── NHÓM 8: HIỂN THỊ APP ───────────────────────────────────────────────────
    {
        "id": "TC076", "module": "Hiển thị App",
        "title": "TK Chờ duyệt: Trang chủ hiển thị SP mới/bán chạy với giá liên hệ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK trạng thái Chờ duyệt",
        "steps": "1. Đăng nhập App\n2. Quan sát Trang chủ",
        "expected": "- Banner quảng bá hiển thị (nếu có, xét theo đối tượng)\n- SP mới/bán chạy thuộc danh mục đã chọn hiển thị với Giá liên hệ",
        "priority": "High"
    },
    {
        "id": "TC077", "module": "Hiển thị App",
        "title": "TK Chờ duyệt: Trang SP hiển thị danh mục đã chọn với giá liên hệ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK Chờ duyệt",
        "steps": "1. Đăng nhập App\n2. Vào Trang Sản phẩm",
        "expected": "- Hiển thị nút đóng/mở Danh mục SP + câu hỏi khảo sát\n- Các danh mục đã chọn hiển thị\n- Giá SP = Giá liên hệ\n- Không xét ẩn/hiện SP theo đối tượng (TK chưa được duyệt)",
        "priority": "High"
    },
    {
        "id": "TC078", "module": "Hiển thị App",
        "title": "TK Chờ duyệt: Trang CT/Thử thách hiển thị nhưng bị mờ",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK Chờ duyệt",
        "steps": "1. Đăng nhập App\n2. Vào Trang CT / Thử thách",
        "expected": "- Trang CT, Thử thách hiển thị nhưng bị MỜ (không tương tác được)",
        "priority": "High"
    },
    {
        "id": "TC079", "module": "Hiển thị App",
        "title": "TK Chờ duyệt: Trang Tài khoản hiển thị nút Đăng xuất",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK Chờ duyệt",
        "steps": "1. Đăng nhập App\n2. Vào Trang Tài khoản",
        "expected": "- Trang Tài khoản hiển thị\n- Có nút Đăng xuất",
        "priority": "Medium"
    },
    {
        "id": "TC080", "module": "Hiển thị App",
        "title": "TK Đang hoạt động: Tất cả Tab hiển thị đầy đủ và rõ ràng",
        "type": "Happy Path",
        "preconditions": "Đăng nhập App bằng TK Đang hoạt động",
        "steps": "1. Đăng nhập App\n2. Quan sát tất cả Tab",
        "expected": "- Tất cả Tab hiển thị RÕ RÀNG và ĐẦY ĐỦ\n- Áp dụng ẩn/hiện SP theo đối tượng KH\n- Giá bán hiển thị đúng (nếu không bị khóa)",
        "priority": "High"
    },
    {
        "id": "TC081", "module": "Hiển thị App",
        "title": "TK Ngưng hoạt động: Không đăng nhập App được",
        "type": "Negative",
        "preconditions": "TK KH bị Tạm dừng (Ngưng hoạt động)",
        "steps": "1. Mở App\n2. Đăng nhập bằng TK Ngưng hoạt động",
        "expected": "- Hiển thị thông báo lỗi: Tài khoản đã bị tạm dừng (hoặc tương tự)\n- Không cho phép đăng nhập vào App",
        "priority": "High"
    },
    {
        "id": "TC082", "module": "Hiển thị App",
        "title": "Đóng/Mở danh mục SP trên App: hiển thị câu hỏi khảo sát đúng",
        "type": "Happy Path",
        "preconditions": "TK đang hoạt động, đang ở Trang Sản phẩm",
        "steps": "1. Nhấn Mở một danh mục mới (chưa được chọn trước đó)\n2. Quan sát bộ câu hỏi khảo sát xuất hiện\n3. Trả lời đủ câu hỏi",
        "expected": "- Danh mục mở thành công, hiển thị trong Left Menu\n- Bộ câu hỏi khảo sát đúng với danh mục cấp 1\n- Đóng danh mục → danh mục biến khỏi Left Menu",
        "priority": "Medium"
    },
    {
        "id": "TC083", "module": "Hiển thị App",
        "title": "Hiển thị thông tin Cấp bậc đúng theo Loại KH (Sỉ vs Lẻ)",
        "type": "Happy Path",
        "preconditions": "Có TK KH Sỉ (cấp Thạch Anh) và TK KH Lẻ",
        "steps": "1. Đăng nhập App bằng TK Sỉ → xem mục Cấp bậc\n2. Đăng xuất → đăng nhập bằng TK Lẻ → xem mục Cấp bậc",
        "expected": "- KH Sỉ: hiển thị cấp bậc hiện tại, chiết khấu, điều kiện cấp tiếp\n- KH Lẻ: hiển thị Thành viên, không có thông tin cấp bậc",
        "priority": "Medium"
    },
]


# ─── HÀM TẠO EXCEL ────────────────────────────────────────────────────────────

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Màu sắc ────────────────────────────────────────────────────────────────
    COLOR_HEADER      = "1F3864"   # Xanh đậm
    COLOR_GROUP_ROW   = "2E75B6"   # Xanh vừa (dòng nhóm)
    COLOR_HAPPY       = "E2EFDA"   # Xanh lá nhạt
    COLOR_NEGATIVE    = "FCE4D6"   # Cam nhạt
    COLOR_BOUNDARY    = "FFF2CC"   # Vàng nhạt
    COLOR_HIGH        = "C00000"   # Đỏ đậm
    COLOR_MEDIUM      = "ED7D31"   # Cam
    COLOR_LOW         = "70AD47"   # Xanh lá

    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_group   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_normal  = Font(name="Calibri", size=9)
    font_id      = Font(name="Calibri", bold=True, size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    align_left_c = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="595959")
    border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ── Tiêu đề tổng ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "BỘ TEST CASE – APP VINAGO (MODULE KHÁCH HÀNG)   |   Tài liệu: 260803_Khach hang_V2.pdf   |   Ngày tạo: 2026-08-11"
    title_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header cột ────────────────────────────────────────────────────────────
    headers = [
        "Test Case ID", "Module", "Title", "Test Type",
        "Preconditions", "Steps", "Expected Result", "Priority"
    ]
    col_widths = [13, 22, 42, 14, 38, 50, 45, 10]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 22

    # ── Nhóm test ──────────────────────────────────────────────────────────────
    groups = [
        ("NHÓM 1: ĐĂNG KÝ TÀI KHOẢN (APP)",                "TC001", "TC019"),
        ("NHÓM 2: QUẢN LÝ TRẠNG THÁI TÀI KHOẢN (CMS)",     "TC020", "TC031"),
        ("NHÓM 3: TÌM KIẾM & BỘ LỌC – DS TÀI KHOẢN",       "TC032", "TC040"),
        ("NHÓM 4: DANH SÁCH NGƯỜI MUA – TẠO MỚI & QUẢN LÝ","TC041", "TC050"),
        ("NHÓM 5: TÀI KHOẢN CHÍNH/PHỤ – PHÂN QUYỀN",        "TC051", "TC060"),
        ("NHÓM 6: HỆ THỐNG CẤP BẬC KHÁCH BÁN SỈ",          "TC061", "TC068"),
        ("NHÓM 7: CHUYỂN ĐỔI LOẠI KHÁCH HÀNG SỈ ↔ LẺ",     "TC069", "TC075"),
        ("NHÓM 8: HIỂN THỊ APP THEO TRẠNG THÁI TÀI KHOẢN",  "TC076", "TC083"),
    ]

    type_color = {
        "Happy Path": COLOR_HAPPY,
        "Negative":   COLOR_NEGATIVE,
        "Boundary":   COLOR_BOUNDARY,
    }
    priority_color = {
        "High":   COLOR_HIGH,
        "Medium": COLOR_MEDIUM,
        "Low":    COLOR_LOW,
    }

    row_cursor = 3
    tc_map = {tc["id"]: tc for tc in test_cases}
    tc_ids = [tc["id"] for tc in test_cases]

    for group_name, start_id, end_id in groups:
        # Dòng nhóm
        ws.merge_cells(f"A{row_cursor}:H{row_cursor}")
        g_cell = ws.cell(row=row_cursor, column=1, value=f"  {group_name}")
        g_cell.font      = font_group
        g_cell.fill      = PatternFill("solid", fgColor=COLOR_GROUP_ROW)
        g_cell.alignment = Alignment(horizontal="left", vertical="center")
        g_cell.border    = border_thick
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        # Dòng test case trong nhóm
        start_idx = tc_ids.index(start_id)
        end_idx   = tc_ids.index(end_id) + 1
        group_tcs = test_cases[start_idx:end_idx]

        for tc in group_tcs:
            bg = type_color.get(tc["type"], "FFFFFF")
            fill = PatternFill("solid", fgColor=bg)

            values = [
                tc["id"], tc["module"], tc["title"], tc["type"],
                tc["preconditions"], tc["steps"],
                tc["expected"], tc["priority"]
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=val)
                cell.fill   = fill
                cell.border = border_all

                if col_idx == 1:  # ID
                    cell.font      = font_id
                    cell.alignment = align_center
                elif col_idx == 4:  # Test Type
                    cell.font      = Font(name="Calibri", bold=True, size=9,
                                          color=("276221" if tc["type"]=="Happy Path"
                                                 else "7B2C00" if tc["type"]=="Negative"
                                                 else "7B6300"))
                    cell.alignment = align_center
                elif col_idx == 8:  # Priority
                    p_color = priority_color.get(tc["priority"], "000000")
                    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
                    cell.fill      = PatternFill("solid", fgColor=p_color)
                    cell.alignment = align_center
                elif col_idx in (2, 3):
                    cell.font      = font_normal
                    cell.alignment = align_left_c
                else:
                    cell.font      = font_normal
                    cell.alignment = align_left

            ws.row_dimensions[row_cursor].height = 90
            row_cursor += 1

        row_cursor += 1  # Dòng trống giữa nhóm

    # ── Sheet tóm tắt ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tổng kết")
    summary_headers = ["Nhóm", "Từ TC", "Đến TC", "Số TC", "Happy Path", "Negative", "Boundary"]
    summary_data = [
        ("Đăng ký Tài khoản (App)",         "TC001","TC019", 19, 8, 7, 4),
        ("Quản lý Trạng thái TK (CMS)",     "TC020","TC031", 12, 9, 1, 2),
        ("Tìm kiếm & Bộ lọc – DS TK",       "TC032","TC040",  9, 8, 1, 0),
        ("DS Người mua – Tạo mới & QL",     "TC041","TC050", 10, 8, 2, 0),
        ("TK Chính/Phụ – Phân quyền",       "TC051","TC060", 10, 7, 2, 1),
        ("Hệ thống Cấp bậc KH Sỉ",         "TC061","TC068",  8, 5, 1, 2),
        ("Chuyển đổi Loại KH Sỉ ↔ Lẻ",     "TC069","TC075",  7, 4, 0, 3),
        ("Hiển thị App theo Trạng thái",    "TC076","TC083",  8, 7, 1, 0),
    ]

    ws2.merge_cells("A1:G1")
    s_title = ws2["A1"]
    s_title.value     = "TỔNG KẾT BỘ TEST CASE – APP VINAGO (MODULE KHÁCH HÀNG)"
    s_title.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    s_title.fill      = PatternFill("solid", fgColor="1F3864")
    s_title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = font_header
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = align_center
        cell.border    = border_all

    s_widths = [42, 10, 10, 8, 12, 10, 10]
    for i, w in enumerate(s_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    totals = [0, 0, 0, 0]
    for r_idx, row in enumerate(summary_data, start=3):
        fills = [PatternFill("solid", fgColor="DEEAF1"), PatternFill("solid", fgColor="FFFFFF")]
        row_fill = fills[r_idx % 2]
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10, bold=(col_idx==1))
            cell.fill      = row_fill
            cell.alignment = align_center if col_idx > 1 else Alignment(horizontal="left", vertical="center")
            cell.border    = border_all
        totals[0] += row[3]; totals[1] += row[4]; totals[2] += row[5]; totals[3] += row[6]

    total_row = r_idx + 1
    ws2.merge_cells(f"A{total_row}:C{total_row}")
    tc = ws2.cell(row=total_row, column=1, value="TỔNG CỘNG")
    tc.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    tc.fill      = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = border_thick

    for col_idx, val in enumerate([totals[0], totals[1], totals[2], totals[3]], start=4):
        cell = ws2.cell(row=total_row, column=col_idx, value=val)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = align_center
        cell.border    = border_thick

    ws2.row_dimensions[total_row].height = 22

    # ── Lưu file ───────────────────────────────────────────────────────────────
    output_path = r"d:\Java lean\TestCase\TestCase_VINAGO_KhachHang.xlsx"
    wb.save(output_path)
    print(f"✅ Đã tạo file Excel thành công: {output_path}")
    print(f"   - Sheet 'Test Cases': {len(test_cases)} test cases")
    print(f"   - Sheet 'Tổng kết': Bảng tóm tắt")


if __name__ == "__main__":
    create_excel()
