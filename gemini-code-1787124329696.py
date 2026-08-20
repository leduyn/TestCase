import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter  # Import hàm get_column_letter

# 1. Dữ liệu Test Cases
test_cases_data = [
    # Module 1
    {"ID": "TC_REG_001", "Module": "Đăng ký (App)", "Title": "Đăng ký tài khoản App thành công (Happy Path)", "Type": "Happy Path", "Preconditions": "App đã được mở", "Steps": "1. Nhập Họ tên, SĐT, Ngày sinh, Giới tính, Tên cửa hàng.\n2. Chọn Tỉnh/Thành, Phường/Xã, Địa chỉ.\n3. Nhập MST, Tên đơn vị, Email (1-5 email), Mật khẩu.\n4. Upload 1-5 ảnh đại diện.\n5. Chọn 1 Danh mục SP & trả lời khảo sát.\n6. Xác nhận đăng ký & Nhập OTP đúng.", "Expected Result": "1. Đăng ký thành công (thông báo 1).\n2. CMS hiển thị 'Chờ duyệt', Loại KH 'Khách bán Lẻ'.\n3. App giới hạn tính năng (Banner, SP giá liên hệ).", "Priority": "High"},
    {"ID": "TC_REG_002", "Module": "Đăng ký (App)", "Title": "Đăng ký thất bại khi bỏ trống trường bắt buộc", "Type": "Negative", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Để trống một trong các trường (*): Họ tên, SĐT, Tên cửa hàng, Tỉnh/Thành, Phường/Xã, Địa chỉ, MST, Tên đơn vị, Email, Mật khẩu.\n2. Bấm Xác nhận.", "Expected Result": "Hiển thị cảnh báo lỗi tương ứng tại từng trường bắt buộc, không cho tiếp tục.", "Priority": "High"},
    {"ID": "TC_REG_003", "Module": "Đăng ký (App)", "Title": "Kiểm tra trùng SĐT khi đăng ký tài khoản chính", "Type": "Negative", "Preconditions": "SĐT A đã tồn tại trong DS tài khoản", "Steps": "1. Nhập SĐT A vào trường SĐT đăng ký.\n2. Nhập đầy đủ các thông tin hợp lệ khác.\n3. Bấm Tiếp tục.", "Expected Result": "Báo lỗi SĐT đã được sử dụng, không cho tiến hành nhận OTP.", "Priority": "High"},
    {"ID": "TC_REG_004", "Module": "Đăng ký (App)", "Title": "Kiểm tra trùng MST trong Danh sách Khách hàng", "Type": "Negative", "Preconditions": "MST B đã tồn tại trong DS Khách hàng", "Steps": "1. Nhập MST B vào thông tin xuất hóa đơn.\n2. Nhập các thông tin khác hợp lệ.\n3. Bấm Đăng ký.", "Expected Result": "Báo lỗi MST đã tồn tại trong hệ thống Khách hàng, không cho tạo tài khoản.", "Priority": "High"},
    {"ID": "TC_REG_005", "Module": "Đăng ký (App)", "Title": "Validation trường Email - Nhập vượt quá 5 email", "Type": "Boundary", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Tại trường Email XHĐ, nhập 6 địa chỉ email phân tách hợp lệ.\n2. Bấm xác nhận.", "Expected Result": "Báo lỗi chỉ cho phép nhập tối đa 5 email.", "Priority": "Medium"},
    {"ID": "TC_REG_006", "Module": "Đăng ký (App)", "Title": "Validation trường Email - Định dạng sai", "Type": "Negative", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Nhập email sai định dạng (abc@com, test@domain,com, abc@@gmail.com).", "Expected Result": "Báo lỗi định dạng email không hợp lệ.", "Priority": "Medium"},
    {"ID": "TC_REG_007", "Module": "Đăng ký (App)", "Title": "Validation Mật khẩu - Độ dài nhỏ hơn 6 ký tự", "Type": "Boundary", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Nhập mật khẩu 5 ký tự (12345).", "Expected Result": "Báo lỗi Mật khẩu phải có tối thiểu 6 ký tự.", "Priority": "High"},
    {"ID": "TC_REG_008", "Module": "Đăng ký (App)", "Title": "Validation Nhập lại mật khẩu không khớp", "Type": "Negative", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Nhập Mật khẩu: 123456.\n2. Nhập Nhập lại mật khẩu: 1234567.", "Expected Result": "Báo lỗi Mật khẩu nhập lại không trùng khớp.", "Priority": "High"},
    {"ID": "TC_REG_009", "Module": "Đăng ký (App)", "Title": "Validation Upload Hình ảnh (Avatar)", "Type": "Boundary", "Preconditions": "Màn hình đăng ký App", "Steps": "1. Tải lên 0 hình ảnh.\n2. Tải lên 6 hình ảnh.", "Expected Result": "1. Báo lỗi yêu cầu ít nhất 1 hình ảnh.\n2. Chặn không cho upload quá 5 hình ảnh, lấy hình đầu tiên làm Avatar.", "Priority": "Medium"},
    {"ID": "TC_REG_010", "Module": "Đăng ký (App)", "Title": "Bắt buộc chọn Danh mục sản phẩm & Khảo sát", "Type": "Negative", "Preconditions": "Màn hình chọn danh mục", "Steps": "1. Không tích chọn danh mục sản phẩm nào.\n2. Bấm Tiếp tục.", "Expected Result": "Báo lỗi bắt buộc chọn ít nhất 1 danh mục sản phẩm cấp 1.", "Priority": "High"},
    {"ID": "TC_REG_011", "Module": "Đăng ký (App)", "Title": "Nhập OTP sai / Hết hạn", "Type": "Negative", "Preconditions": "Màn hình nhập OTP", "Steps": "1. Nhập OTP sai 3 lần hoặc nhập OTP khi đã quá thời gian hết hạn.", "Expected Result": "Báo lỗi OTP không chính xác/hết hạn, không khởi tạo tài khoản.", "Priority": "High"},
    {"ID": "TC_REG_012", "Module": "Tạo mới (CMS)", "Title": "Admin tạo mới Khách hàng từ CMS (Happy Path)", "Type": "Happy Path", "Preconditions": "Đã đăng nhập CMS Admin", "Steps": "1. Vào Module Khách hàng > Tạo mới.\n2. Chọn Loại KH (Khách bán sỉ / Khách bán lẻ).\n3. Nhập đầy đủ thông tin bắt buộc & Thông tin XHĐ.\n4. Bấm Lưu.", "Expected Result": "Tạo thành công khách hàng ở trạng thái 'Chờ duyệt'.", "Priority": "High"},
    {"ID": "TC_REG_013", "Module": "Bảo mật", "Title": "Kiểm tra SQL Injection & XSS vào các trường văn bản", "Type": "Security", "Preconditions": "Màn hình Đăng ký / Tạo mới", "Steps": "1. Nhập chuỗi <script>alert('XSS')</script> hoặc ' OR '1'='1 vào Họ tên, Tên cửa hàng, Địa chỉ.", "Expected Result": "Hệ thống sanitize dữ liệu, lưu dưới dạng plain text, không thực thi script hoặc lỗi SQL.", "Priority": "High"},

    # Module 2
    {"ID": "TC_CUST_001", "Module": "DS Khách hàng", "Title": "Duyệt tài khoản Khách hàng (Chờ duyệt -> Đang hoạt động)", "Type": "Happy Path", "Preconditions": "KH ở trạng thái 'Chờ duyệt'", "Steps": "1. Mở chi tiết KH 'Chờ duyệt'.\n2. Bấm 'Duyệt'.", "Expected Result": "1. Trạng thái KH chuyển sang 'Đang hoạt động'.\n2. Tự sinh Mã KH, Nickname.\n3. Tạo mới/Cập nhật thông tin vào DS Người mua.\n4. App KH nhận thông báo thành công (1), mở đầy đủ tính năng.", "Priority": "High"},
    {"ID": "TC_CUST_002", "Module": "DS Khách hàng", "Title": "Kiểm tra cú pháp sinh Mã KH khi duyệt", "Type": "Boundary", "Preconditions": "Tài khoản được duyệt", "Steps": "Duyệt tài khoản có Tỉnh cũ: Trà Vinh (TVH), Tỉnh mới: Vĩnh Long (VLG), Năm 2026, Cấp bậc: Thạch Anh 1.", "Expected Result": "Mã KH sinh ra đúng dạng: TVH[STT 3 chữ số]VLG26_TA1 (Ví dụ: TVH002VLG26_TA1).", "Priority": "High"},
    {"ID": "TC_CUST_003", "Module": "DS Khách hàng", "Title": "Từ chối tài khoản Khách hàng (Chờ duyệt -> Từ chối)", "Type": "Happy Path", "Preconditions": "KH ở trạng thái 'Chờ duyệt'", "Steps": "1. Mở chi tiết KH 'Chờ duyệt'.\n2. Bấm 'Từ chối'.", "Expected Result": "1. Trạng thái chuyển 'Từ chối'.\n2. App nhận thông báo thất bại (2).\n3. Hệ thống KHÔNG kiểm tra trùng thông tin với các tài khoản Từ chối này.", "Priority": "High"},
    {"ID": "TC_CUST_004", "Module": "DS Khách hàng", "Title": "Kiểm tra quy tắc sinh và sửa Nickname", "Type": "Boundary", "Preconditions": "KH đã được duyệt", "Steps": "1. Kiểm tra Nickname tự sinh.\n2. Vào chỉnh sửa Nickname thành chuỗi 15 ký tự.\n3. Chỉnh sửa Nickname chứa ký tự đặc biệt (#$@) hoặc >15 ký tự.", "Expected Result": "1. Nickname tự sinh dạng D0001 (STT tăng dần).\n2. Lưu thành công 15 ký tự hợp lệ.\n3. Báo lỗi khi chứa ký tự đặc biệt hoặc >15 ký tự.", "Priority": "Medium"},
    {"ID": "TC_CUST_005", "Module": "DS Khách hàng", "Title": "Chuyển loại KH: Khách bán Lẻ -> Khách bán Sỉ", "Type": "Logic", "Preconditions": "KH đang là Khách bán Lẻ", "Steps": "1. Đổi Loại KH sang 'Khách bán Sỉ'.\n2. Lưu thay đổi.", "Expected Result": "1. Cập nhật Cấp bậc ngay thành 'Thạch Anh' (tối thiểu).\n2. Ghi nhận lịch sử thay đổi Loại KH.", "Priority": "High"},
    {"ID": "TC_CUST_006", "Module": "DS Khách hàng", "Title": "Chuyển loại KH: Khách bán Sỉ -> Khách bán Lẻ (Xử lý Phí bảo lãnh)", "Type": "Logic", "Preconditions": "Khách bán Sỉ còn dư phí bảo lãnh và đơn hàng", "Steps": "1. Chuyển KH Sỉ sang KH Lẻ.\n2. Xác nhận.", "Expected Result": "1. Cấn trừ phí bảo lãnh cho đơn chưa thanh toán.\n2. Phí dư chuyển về công nợ cuối kỳ.\n3. Cập nhật cấp bậc về 'Thành viên'.", "Priority": "High"},
    {"ID": "TC_CUST_007", "Module": "DS Khách hàng", "Title": "Kiểm tra Khách ngưng hoạt động truy cập App", "Type": "Negative", "Preconditions": "KH có trạng thái 'Ngưng hoạt động'", "Steps": "Đăng nhập tài khoản KH Ngưng hoạt động trên App.", "Expected Result": "Không đăng nhập/không đặt hàng được, hiển thị thông báo tài khoản ngưng hoạt động.", "Priority": "High"},
    {"ID": "TC_CUST_008", "Module": "DS Khách hàng", "Title": "Kiểm tra bộ lọc và tìm kiếm DS Khách hàng", "Type": "Functional", "Preconditions": "Có dữ liệu KH đa dạng", "Steps": "Tìm kiếm theo Mã KH, Tên KH, MST, SĐT đăng nhập; Lọc theo Loại KH, Cấp bậc, Vùng KD (V1-V8), Trạng thái.", "Expected Result": "Trả về đúng danh sách dữ liệu khớp với điều kiện lọc.", "Priority": "Medium"},
    {"ID": "TC_CUST_009", "Module": "DS Khách hàng", "Title": "Xuất dữ liệu Excel Khách hàng", "Type": "Functional", "Preconditions": "Có dữ liệu DS KH", "Steps": "Bấm 'Xuất dữ liệu theo bộ lọc'.", "Expected Result": "Tải về file Excel chứa đủ các cột thông tin quy định.", "Priority": "Low"},
    {"ID": "TC_CUST_010", "Module": "Khảo sát (App)", "Title": "Kiểm tra quy tắc khảo sát khi mở Danh mục sản phẩm", "Type": "Logic", "Preconditions": "Đã đăng nhập TKC/TKP", "Steps": "1. Mở danh mục SP chưa từng trả lời khảo sát.\n2. Đóng/mở lại danh mục đó.", "Expected Result": "1. Hiển thị bộ câu hỏi tương ứng.\n2. Lần đóng/mở sau không hiển thị câu hỏi khảo sát nữa (chỉ trả lời 1 lần).", "Priority": "Medium"},

    # Module 3
    {"ID": "TC_BUYER_001", "Module": "DS Người mua", "Title": "Tạo trực tiếp Người mua từ CMS - MST mới", "Type": "Happy Path", "Preconditions": "Màn hình DS Người mua", "Steps": "1. Bấm Tạo mới.\n2. Nhập MST mới, Tên đơn vị, Địa chỉ, Email (1-5 email).\n3. Lưu.", "Expected Result": "1. Tạo thành công Người mua ở trạng thái 'Nháp'.\n2. Tự sinh Mã người mua cú pháp NMxxxx (STT tăng dần).", "Priority": "High"},
    {"ID": "TC_BUYER_002", "Module": "DS Người mua", "Title": "Tạo trực tiếp Người mua từ CMS - Trùng MST", "Type": "Boundary", "Preconditions": "MST C đã có trong DS Người mua", "Steps": "1. Nhập MST C.\n2. Bấm Lưu.", "Expected Result": "Hiển thị cảnh báo: 'Trùng MST trong DS người mua. Vẫn muốn tạo mới thông tin người mua? Hủy bỏ | Xác nhận'.", "Priority": "High"},
    {"ID": "TC_BUYER_003", "Module": "DS Người mua", "Title": "Chuyển trạng thái Nháp -> Chờ xác nhận", "Type": "Logic", "Preconditions": "Người mua ở trạng thái 'Nháp'", "Steps": "Bấm Chuyển trạng thái sang 'Chờ xác nhận'.", "Expected Result": "1. Hệ thống kiểm tra lại MST.\n2. Truyền dữ liệu sang Bravo (Mã NM, MST, Tên đơn vị, Địa chỉ XHĐ, Email, Mã KH, Tên KH).", "Priority": "High"},
    {"ID": "TC_BUYER_004", "Module": "DS Người mua", "Title": "Tự động tạo Người mua từ Đơn hàng App", "Type": "Logic", "Preconditions": "Đơn hàng App có MST mới", "Steps": "Chuyển đơn hàng sang 'Soạn hàng' mà MST chưa có trong DS Người mua.", "Expected Result": "Hệ thống tự động tạo Người mua mới trạng thái 'Chờ xác nhận' và truyền sang Bravo.", "Priority": "High"},
    {"ID": "TC_BUYER_005", "Module": "DS Người mua", "Title": "Duyệt tài khoản App - MST trùng trong DS Người mua (Chọn Cập nhật)", "Type": "Logic", "Preconditions": "Tài khoản App có MST đã tồn tại", "Steps": "1. Duyệt tài khoản App.\n2. Chọn 'Cập nhật' khi bị trùng MST.", "Expected Result": "1. Cập nhật Mã KH, Tên KH, Loại KH, Trạng thái KH vào NM cũ.\n2. Giữ nguyên Tên đơn vị, Địa chỉ, Email.\n3. Truyền Mã KH, Tên KH sang Bravo.", "Priority": "High"},
    {"ID": "TC_BUYER_006", "Module": "DS Người mua", "Title": "Duyệt tài khoản App - MST trùng trong DS Người mua (Chọn Tạo mới)", "Type": "Logic", "Preconditions": "Tài khoản App có MST đã tồn tại", "Steps": "1. Duyệt tài khoản App.\n2. Chọn 'Tạo mới' khi bị trùng MST.", "Expected Result": "Sinh ra một Mã người mua mới (NMxxxx) theo thông tin XHĐ của Khách hàng.", "Priority": "High"},
    {"ID": "TC_BUYER_007", "Module": "Bravo Sync", "Title": "Đồng bộ luồng Cập nhật Người mua từ Bravo về CMS", "Type": "Integration", "Preconditions": "Trạng thái NM = Đang giao dịch / Ngưng giao dịch", "Steps": "Bravo truyền điều chỉnh (MST, Tên đơn vị, Địa chỉ XHĐ, Email) sang CMS.", "Expected Result": "CMS cập nhật thông tin XHĐ theo dữ liệu Bravo truyền về.", "Priority": "High"},
    {"ID": "TC_BUYER_008", "Module": "Bravo Sync", "Title": "Luồng Hủy Người mua từ Bravo", "Type": "Integration", "Preconditions": "Người mua đang hoạt động", "Steps": "Bravo trả trạng thái 'Hủy' về CMS.", "Expected Result": "1. Trạng thái NM trên CMS chuyển thành 'Hủy'.\n2. CMS ngắt nhận dữ liệu từ Bravo cho bản ghi này.\n3. Không kiểm tra trùng MST với NM trạng thái Hủy.", "Priority": "High"},
    {"ID": "TC_BUYER_009", "Module": "DS Người mua", "Title": "Đặt hàng với Người mua có trạng thái 'Ngưng giao dịch'", "Type": "Negative", "Preconditions": "Người mua ở trạng thái Ngưng giao dịch", "Steps": "Tạo đơn hàng gắn với Người mua có trạng thái 'Ngưng giao dịch'.", "Expected Result": "Hệ thống cảnh báo: 'Không thực hiện giao dịch đối với MST này', không cho tạo đơn.", "Priority": "High"},

    # Module 4
    {"ID": "TC_ADDR_001", "Module": "Địa chỉ giao hàng", "Title": "Tạo mới Địa chỉ giao hàng thủ công (Happy Path)", "Type": "Happy Path", "Preconditions": "Màn hình DS Địa chỉ giao hàng", "Steps": "1. Bấm Tạo mới.\n2. Nhập Tên địa chỉ, Địa chỉ, Người liên hệ, SĐT, Ghi chú.\n3. Chọn Hình thức giao (Chành/Trực tiếp/Kho).\n4. Chọn/nhập Mã người mua (*).\n5. Bấm Lưu.", "Expected Result": "1. Tạo thành công, tự sinh Mã giao hàng GHxxxx.\n2. Dữ liệu truyền đầy đủ sang Bravo.", "Priority": "High"},
    {"ID": "TC_ADDR_002", "Module": "Địa chỉ giao hàng", "Title": "Tạo mới Địa chỉ giao hàng - Mã người mua không tồn tại", "Type": "Negative", "Preconditions": "Màn hình Tạo mới địa chỉ", "Steps": "1. Nhập Mã người mua NM999999 (chưa có trong hệ thống).\n2. Bấm Lưu.", "Expected Result": "Báo lỗi: 'Mã Người mua chưa tồn tại trên hệ thống', không cho lưu.", "Priority": "High"},
    {"ID": "TC_ADDR_003", "Module": "Địa chỉ giao hàng", "Title": "Import file Excel Địa chỉ giao hàng thành công", "Type": "Functional", "Preconditions": "File Excel đúng mẫu", "Steps": "1. Bấm Import file.\n2. Chọn file Excel chứa 10 dòng địa chỉ hợp lệ.\n3. Bấm Tải lên.", "Expected Result": "Import thành công 10 địa chỉ, tự sinh 10 Mã giao hàng tương ứng và gán đúng Mã người mua.", "Priority": "Medium"},
    {"ID": "TC_ADDR_004", "Module": "Địa chỉ giao hàng", "Title": "Import file Excel - File chứa Mã người mua lỗi", "Type": "Negative", "Preconditions": "File Excel mẫu có 1 dòng sai Mã NM", "Steps": "1. Tải lên file Excel có dòng thứ 3 chứa Mã NM không tồn tại.\n2. Bấm Tải lên.", "Expected Result": "Báo lỗi chi tiết tại dòng số 3, chặn import dòng lỗi.", "Priority": "Medium"},
    {"ID": "TC_ADDR_005", "Module": "Địa chỉ giao hàng", "Title": "Tự động chọn Địa chỉ giao hàng mặc định", "Type": "Logic", "Preconditions": "Người mua có nhiều địa chỉ giao hàng", "Steps": "Khách hàng thực hiện tạo đơn hàng mới trên App.", "Expected Result": "Tự động lấy Địa chỉ nhận hàng mà Khách hàng đã sử dụng gần nhất làm địa chỉ mặc định.", "Priority": "Low"},

    # Module 5
    {"ID": "TC_SUB_001", "Module": "Phân quyền TK", "Title": "TKC tạo Tài khoản phụ thành công", "Type": "Happy Path", "Preconditions": "Đăng nhập TKC trên App", "Steps": "1. Vào Quản lý TK phụ > Tạo mới.\n2. Nhập Tên TK, SĐT chưa đăng ký, Mật khẩu.\n3. Tích chọn/Không tích chọn các quyền chặn.\n4. Bấm Tạo.", "Expected Result": "1. Tạo thành công TKP ở trạng thái 'Chờ kích hoạt'.\n2. SĐT được kiểm tra trùng trong hệ thống.", "Priority": "High"},
    {"ID": "TC_SUB_002", "Module": "Phân quyền TK", "Title": "Thứ tự hiển thị Danh sách Tài khoản phụ", "Type": "Logic", "Preconditions": "Có nhiều TKP ở các trạng thái khác nhau", "Steps": "Xem DS Tài khoản phụ trên App/CMS.", "Expected Result": "Ưu tiên Trạng thái: Chờ kích hoạt > Đang hoạt động > Khóa > Từ chối. Nếu cùng trạng thái, ưu tiên thời gian tạo từ gần đến xa.", "Priority": "Medium"},
    {"ID": "TC_SUB_003", "Module": "Phân quyền TK", "Title": "Đăng nhập TKP ở trạng thái 'Chờ kích hoạt'", "Type": "Negative", "Preconditions": "TKP ở trạng thái Chờ kích hoạt", "Steps": "Đăng nhập SĐT/Mật khẩu của TKP trên App.", "Expected Result": "Báo lỗi tài khoản chưa được kích hoạt, không cho phép truy cập.", "Priority": "High"},
    {"ID": "TC_SUB_004", "Module": "Phân quyền TK", "Title": "CMS Kích hoạt TKP (Chờ kích hoạt -> Đang hoạt động)", "Type": "Happy Path", "Preconditions": "TKP đang Chờ kích hoạt", "Steps": "CMS Admin bấm 'Kích hoạt TK'.", "Expected Result": "1. TKP chuyển sang trạng thái 'Đang hoạt động'.\n2. TKP nhận OTP về SĐT để hoàn tất đăng nhập.", "Priority": "High"},
    {"ID": "TC_SUB_005", "Module": "Phân quyền TK", "Title": "Kiểm tra Phân quyền: 'Không thấy giá bán'", "Type": "Security", "Preconditions": "TKP bị chặn quyền Giá bán", "Steps": "Đăng nhập TKP > Mở trang Sản phẩm / Theo dõi đơn hàng.", "Expected Result": "1. Ẩn toàn bộ giá bán sản phẩm.\n2. Tại Danh sách theo dõi đơn hàng: Không cho xem chi tiết đơn hàng.", "Priority": "High"},
    {"ID": "TC_SUB_006", "Module": "Phân quyền TK", "Title": "Kiểm tra Phân quyền: 'Không cho phép Đặt hàng'", "Type": "Functional", "Preconditions": "TKP bị chặn quyền Đặt hàng", "Steps": "Đăng nhập TKP > Vào xem Sản phẩm / Giỏ hàng / Đơn hàng.", "Expected Result": "1. Mờ/Ẩn nút 'Thêm vào xe hàng', icon thêm nhanh, ẩn icon Xe hàng.\n2. Không thấy nút 'Hủy bỏ' / 'Đặt lại đơn hàng'.", "Priority": "High"},
    {"ID": "TC_SUB_007", "Module": "Phân quyền TK", "Title": "Kiểm tra Phân quyền: 'Không thấy thông tin người mua'", "Type": "Security", "Preconditions": "TKP bị chặn quyền Người mua", "Steps": "Đăng nhập TKP > Mở xem danh sách theo dõi đơn hàng.", "Expected Result": "Không cho phép xem chi tiết đơn hàng và thông tin người mua trên đơn.", "Priority": "High"},
    {"ID": "TC_SUB_008", "Module": "Phân quyền TK", "Title": "Đặt hàng nhập chênh lệch giá & kỳ hạn nợ", "Type": "Functional", "Preconditions": "TKP có quyền Đặt hàng", "Steps": "Đăng nhập TKP > Tiến hành tạo đơn hàng > Nhập chênh lệch giá bán & kỳ hạn nợ.", "Expected Result": "Hệ thống cho phép nhập chênh lệch giá và kỳ hạn nợ bình thường.", "Priority": "Medium"},
    {"ID": "TC_SUB_009", "Module": "Phân quyền TK", "Title": "Chuyển Tài khoản phụ thành Tài khoản chính", "Type": "Logic", "Preconditions": "TKP đang hoạt động", "Steps": "Trên CMS, chọn TKP và bấm 'Chuyển thành Tài khoản chính'.", "Expected Result": "1. TKP chuyển thành TKC.\n2. Xóa bỏ toàn bộ Danh sách phân quyền chặn.", "Priority": "High"},
    {"ID": "TC_SUB_010", "Module": "Phân quyền TK", "Title": "Chuyển Tài khoản chính thành Tài khoản phụ", "Type": "Logic", "Preconditions": "TKC đang hoạt động", "Steps": "Trên CMS, chọn TKC và bấm 'Chuyển thành Tài khoản phụ'.", "Expected Result": "Hiển thị DS phân quyền, mặc định cho phép đầy đủ quyền.", "Priority": "High"},

    # Module 6
    {"ID": "TC_RANK_001", "Module": "Cấp bậc", "Title": "Khách bán Lẻ mặc định Cấp bậc", "Type": "Logic", "Preconditions": "KH là Khách bán Lẻ", "Steps": "Kiểm tra Cấp bậc của KH Bán Lẻ trên App/CMS.", "Expected Result": "Luôn là 'Chưa xếp hạng' (CXH), không áp dụng thăng cấp đá quý.", "Priority": "Medium"},
    {"ID": "TC_RANK_002", "Module": "Cấp bậc", "Title": "Thăng cấp bậc cho Khách bán Sỉ", "Type": "Logic", "Preconditions": "KH Sỉ đạt đủ DTT, Tiền thu, SL Người mua hợp lệ", "Steps": "Chờ đến thời điểm quét tự động (23h ngày cuối tháng).", "Expected Result": "Hệ thống tự động nâng Cấp bậc (VD: Thạch Anh I -> Thạch Anh II) và cập nhật mức chiết khấu.", "Priority": "High"},
    {"ID": "TC_RANK_003", "Module": "Cấp bậc", "Title": "Giảm cấp bậc cho Khách bán Sỉ", "Type": "Logic", "Preconditions": "KH Sỉ không đạt DTT/Tiền thu trong [n] tháng", "Steps": "Chờ đến 23h ngày cuối tháng.", "Expected Result": "Hệ thống tự động giảm Cấp bậc tương ứng.", "Priority": "High"},
    {"ID": "TC_RANK_004", "Module": "Cấp bậc", "Title": "Hiển thị thông tin Cấp bậc trên App KH", "Type": "UI", "Preconditions": "KH Sỉ đã đăng nhập App", "Steps": "Vào màn hình Cấp bậc trên App.", "Expected Result": "Hiển thị chính xác: Cấp bậc hiện tại, mức chiết khấu, Cấp bậc kế tiếp và chỉ số còn thiếu.", "Priority": "Medium"},
    {"ID": "TC_RANK_005", "Module": "Cấp bậc", "Title": "Xử lý dữ liệu Cấp bậc khi qua năm mới", "Type": "Boundary", "Preconditions": "Thời điểm 23:59:59 ngày 31/12", "Steps": "Chờ hệ thống chuyển sang năm mới.", "Expected Result": "1. KHÔNG reset dữ liệu cấp bậc.\n2. Cập nhật lại bộ lọc công nợ theo năm hiện tại.", "Priority": "High"},
]

# 2. Tạo DataFrame
df = pd.DataFrame(test_cases_data)

# 3. Xuất ra Excel với Styling
file_name = "VINAGO_TestCases_Module_KhachHang.xlsx"

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Test Cases', index=False)
    
    # Lấy workbook và worksheet
    workbook = writer.book
    worksheet = writer.sheets['Test Cases']
    
    # Định dạng Font & Header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    
    # Viền (Border)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Format Headers
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format Data Rows
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            # Cấn lề & Auto Wrap Text
            if col_num in [1, 4, 8]:  # ID, Type, Priority
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Format màu cho Priority
            if cell.value == "High":
                cell.font = Font(name="Arial", size=10, bold=True, color="C00000") # Đỏ
            elif cell.value == "Medium":
                cell.font = Font(name="Arial", size=10, color="ED7D31") # Cam

    # -------------------------------------------------------------
    # SỬ DỤNG HÀM get_column_letter ĐỂ TỰ ĐỘNG GÁN ĐỘ RỘNG CỘT
    # -------------------------------------------------------------
    # Mảng lưu kích thước tương ứng theo chỉ số cột (1-indexed)
    widths = [16, 18, 35, 15, 25, 45, 40, 12]
    
    for col_idx, width in enumerate(widths, start=1):
        col_letter = get_column_letter(col_idx)  # Chuyển đổi chỉ số (1, 2, 3...) thành chữ cái (A, B, C...)
        worksheet.column_dimensions[col_letter].width = width

print(f"✅ Đã xuất thành công file Excel sử dụng get_column_letter: {file_name}")